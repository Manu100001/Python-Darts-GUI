#!/usr/bin/python3
# pylint: disable=C0302
"""
This script generates a GUI for the around-the-clock darts game
Single, Double and Triple fields are possible

:author: Manuel Milde manuelmilde@gmx.net
:copyright: 2021 Manuel Milde
"""
import os
from tkinter import Tk
from tkinter import Toplevel
from tkinter import Label
from tkinter import Button
from tkinter import messagebox
from tkinter import Entry
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# color for Excel
greenFill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# months for creating time-based Excel
months = ['Januar', 'Februar', 'März', 'April', 'Mai', 'Juni', 'Juli', 'August', 'September',
          'Oktober', 'November', 'Dezember']


def check_directories():
    """

    :return:
    """
    # create new directories if they do not exist yet
    if not os.path.isdir("Spielstände"):
        os.mkdir("Spielstände")

    if not os.path.isdir("Spielstände/Around-the-clock"):
        os.mkdir("Spielstände/Around-the-clock")

    # current year
    current_year = datetime.now().strftime('%Y')
    if not os.path.isdir("Spielstände/Around-the-clock/" + current_year):
        os.mkdir("Spielstände/Around-the-clock/" + current_year)

    # current month
    current_month = datetime.now().strftime('%m')
    month_name = months[int(current_month) - 1]
    if not os.path.isdir("Spielstände/Around-the-clock/" + current_year + "/" + month_name):
        os.mkdir("Spielstände/Around-the-clock/" + current_year + "/" + month_name)

    # current day
    current_day = int(datetime.now().strftime('%d'))
    date = str(current_day) + "." + str(current_month)
    if not os.path.isdir("Spielstände/Around-the-clock/" + current_year +
                         "/" + month_name + "/" + date):
        os.mkdir("Spielstände/Around-the-clock/" + current_year + "/" + month_name + "/" + date)

    # create new score - file
    time = datetime.now().strftime('%H-%M-%S')

    path = "Spielstände/Around-the-clock/" + current_year + \
           "/" + month_name + "/" + date + "/" + time + ".xlsx"

    return path


def set_standards_in_excel(sheet):
    """

    :return:
    """
    sheet['B5'].fill = greenFill
    sheet['C5'].fill = redFill

    sheet['E5'].fill = greenFill
    sheet['F5'].fill = redFill

    sheet['B6'].fill = greenFill
    sheet['B7'].fill = greenFill
    sheet['B8'].fill = greenFill
    sheet['B9'].fill = greenFill
    sheet['B10'].fill = greenFill
    sheet['B11'].fill = greenFill
    sheet['B12'].fill = greenFill
    sheet['B13'].fill = greenFill
    sheet['B14'].fill = greenFill
    sheet['B15'].fill = greenFill

    sheet['E6'].fill = greenFill
    sheet['E7'].fill = greenFill
    sheet['E8'].fill = greenFill
    sheet['E9'].fill = greenFill
    sheet['E10'].fill = greenFill
    sheet['E11'].fill = greenFill
    sheet['E12'].fill = greenFill
    sheet['E13'].fill = greenFill
    sheet['E14'].fill = greenFill
    sheet['E15'].fill = greenFill

    sheet['L5'].fill = greenFill
    sheet['M5'].fill = redFill
    sheet['N5'].fill = greenFill

    sheet.column_dimensions['L'].width = 9
    sheet.column_dimensions['M'].width = 13


def set_values_in_excel(sheet):
    """

    :return:
    """
    sheet.cell(row=1, column=1).value = "Mode:"

    sheet.cell(row=6, column=2).value = "1"
    sheet.cell(row=7, column=2).value = "2"
    sheet.cell(row=8, column=2).value = "3"
    sheet.cell(row=9, column=2).value = "4"
    sheet.cell(row=10, column=2).value = "5"
    sheet.cell(row=11, column=2).value = "6"
    sheet.cell(row=12, column=2).value = "7"
    sheet.cell(row=13, column=2).value = "8"
    sheet.cell(row=14, column=2).value = "9"
    sheet.cell(row=15, column=2).value = "10"

    sheet.cell(row=6, column=5).value = "11"
    sheet.cell(row=7, column=5).value = "12"
    sheet.cell(row=8, column=5).value = "13"
    sheet.cell(row=9, column=5).value = "14"
    sheet.cell(row=10, column=5).value = "15"
    sheet.cell(row=11, column=5).value = "16"
    sheet.cell(row=12, column=5).value = "17"
    sheet.cell(row=13, column=5).value = "18"
    sheet.cell(row=14, column=5).value = "19"
    sheet.cell(row=15, column=5).value = "20"

    sheet.cell(row=5, column=12).value = "Darts hit"
    sheet.cell(row=5, column=13).value = "Possible darts"
    sheet.cell(row=5, column=14).value = "In percent"

    # show data in Excel
    sheet.cell(row=6, column=3).value = label_count_1['text']
    sheet.cell(row=7, column=3).value = label_count_2['text']
    sheet.cell(row=8, column=3).value = label_count_3['text']
    sheet.cell(row=9, column=3).value = label_count_4['text']
    sheet.cell(row=10, column=3).value = label_count_5['text']
    sheet.cell(row=11, column=3).value = label_count_6['text']
    sheet.cell(row=12, column=3).value = label_count_7['text']
    sheet.cell(row=13, column=3).value = label_count_8['text']
    sheet.cell(row=14, column=3).value = label_count_9['text']
    sheet.cell(row=15, column=3).value = label_count_10['text']

    sheet.cell(row=6, column=6).value = label_count_11['text']
    sheet.cell(row=7, column=6).value = label_count_12['text']
    sheet.cell(row=8, column=6).value = label_count_13['text']
    sheet.cell(row=9, column=6).value = label_count_14['text']
    sheet.cell(row=10, column=6).value = label_count_15['text']
    sheet.cell(row=11, column=6).value = label_count_16['text']
    sheet.cell(row=12, column=6).value = label_count_17['text']
    sheet.cell(row=13, column=6).value = label_count_18['text']
    sheet.cell(row=14, column=6).value = label_count_19['text']
    sheet.cell(row=15, column=6).value = label_count_20['text']


def create_excel():
    """
    This function creates an Excel file in which everything important is listed
    :return:
    """
    path = check_directories()

    excel_file = Workbook()
    sheet = excel_file.create_sheet('Around-the-clock')

    # set cells in excel sheet
    set_standards_in_excel(sheet)
    set_values_in_excel(sheet)

    darts = 0
    all_darts = 0

    if label_single['bg'] == "yellow":
        sheet.cell(row=1, column=2).value = "Single"
        all_darts += 3
    elif label_double['bg'] == "yellow":
        sheet.cell(row=1, column=2).value = "Double"
        all_darts += 3
    else:
        sheet.cell(row=1, column=2).value = "Triple"

    sheet.cell(row=3, column=2).value = "Name:"
    sheet.cell(row=3, column=3).value = input_name1.get()

    sheet.cell(row=5, column=2).value = "Field"
    sheet.cell(row=5, column=3).value = "Hits"

    sheet.cell(row=5, column=5).value = "Field"
    sheet.cell(row=5, column=6).value = "Hits"

    if label_single['bg'] == "yellow":
        sheet.cell(row=5, column=8).value = "Field"
        sheet.cell(row=5, column=9).value = "Hits"
        sheet['H5'].fill = greenFill
        sheet['I5'].fill = redFill
        sheet['H6'].fill = greenFill
        sheet.cell(row=6, column=8).value = "25"
        sheet.cell(row=6, column=9).value = label_count_25['text']
        darts += int(label_count_25['text'])

    if label_double['bg'] == "yellow":
        sheet.cell(row=5, column=8).value = "Field"
        sheet.cell(row=5, column=9).value = "Hits"
        sheet['H5'].fill = greenFill
        sheet['I5'].fill = greenFill
        sheet['H6'].fill = greenFill
        sheet.cell(row=6, column=8).value = "50"
        sheet.cell(row=6, column=9).value = label_count_50['text']
        darts += int(label_count_50['text'])

    sum_darts = get_sum()
    sum_darts += darts

    sheet.cell(row=6, column=12).value = sum_darts
    sheet.cell(row=6, column=13).value = all_darts + 60
    sheet.cell(row=6, column=14).value = round((sum_darts / (all_darts + 60)) * 100, 2)

    # save excel - file
    excel_file.save(path)


def get_sum():
    """
    This function gets the amount of darts of each field
    :return:
    """
    sum_darts = int(label_count_1['text']) + int(label_count_2['text']) + \
        int(label_count_3['text']) + int(label_count_4['text']) + \
        int(label_count_5['text']) + int(label_count_6['text']) + \
        int(label_count_7['text']) + int(label_count_8['text']) + \
        int(label_count_9['text']) + int(label_count_10['text']) + \
        int(label_count_11['text']) + int(label_count_12['text']) + \
        int(label_count_13['text']) + int(label_count_14['text']) + \
        int(label_count_15['text']) + int(label_count_16['text']) + \
        int(label_count_17['text']) + int(label_count_18['text']) + \
        int(label_count_19['text']) + int(label_count_20['text'])

    return sum_darts


def button_exit():
    """
    This function creates an exit - button for the gui
    :return:
    """
    if not any(isinstance(window, Toplevel) for window in gui.winfo_children()):
        exit_window = Toplevel(gui)
        exit_window.geometry('250x150')
        exit_window.resizable(width=False, height=False)
        exit_window.title("Stop?")

        label_exit = Label(exit_window, text="Stop game?", font=('Arial', 11))
        button_yes = Button(exit_window, text="Yes", command=exit_window.quit,
                            font=('Arial', 10, 'bold'), bg="white",
                            fg="green")
        button_no = Button(exit_window, text="No", command=exit_window.destroy,
                           font=('Arial', 10, 'bold'), bg="white", fg="red")

        label_exit.place(x=80, y=0, width=100, height=50)
        button_yes.place(x=50, y=60, width=50, height=50)
        button_no.place(x=150, y=60, width=50, height=50)

    else:
        messagebox.showinfo("Info", "You already clicked on \"Stop\"!")


def reset():
    """
    This function restores the original state
    :return:
    """
    label_selected_mode.pack()
    label_selected_mode.pack_forget()

    label_count_1['bg'] = "yellow"
    label_count_2['bg'] = "white"
    label_count_3['bg'] = "white"
    label_count_4['bg'] = "white"
    label_count_5['bg'] = "white"
    label_count_6['bg'] = "white"
    label_count_7['bg'] = "white"
    label_count_8['bg'] = "white"
    label_count_9['bg'] = "white"
    label_count_10['bg'] = "white"
    label_count_11['bg'] = "white"
    label_count_12['bg'] = "white"
    label_count_13['bg'] = "white"
    label_count_14['bg'] = "white"
    label_count_15['bg'] = "white"
    label_count_16['bg'] = "white"
    label_count_17['bg'] = "white"
    label_count_18['bg'] = "white"
    label_count_19['bg'] = "white"
    label_count_20['bg'] = "white"

    label_count_25['bg'] = "white"
    label_count_50['bg'] = "white"

    label_count_1['text'] = 0
    label_count_2['text'] = 0
    label_count_3['text'] = 0
    label_count_4['text'] = 0
    label_count_5['text'] = 0
    label_count_6['text'] = 0
    label_count_7['text'] = 0
    label_count_8['text'] = 0
    label_count_9['text'] = 0
    label_count_10['text'] = 0
    label_count_11['text'] = 0
    label_count_12['text'] = 0
    label_count_13['text'] = 0
    label_count_14['text'] = 0
    label_count_15['text'] = 0
    label_count_16['text'] = 0
    label_count_17['text'] = 0
    label_count_18['text'] = 0
    label_count_19['text'] = 0
    label_count_20['text'] = 0
    label_count_25['text'] = 0
    label_count_50['text'] = 0

    reset2()


def reset2():
    """
    This function restores the original state
    :return:
    """
    label_count_1.pack()
    label_count_2.pack()
    label_count_3.pack()
    label_count_4.pack()
    label_count_5.pack()
    label_count_6.pack()
    label_count_7.pack()
    label_count_8.pack()
    label_count_9.pack()
    label_count_10.pack()
    label_count_11.pack()
    label_count_12.pack()
    label_count_13.pack()
    label_count_14.pack()
    label_count_15.pack()
    label_count_16.pack()
    label_count_17.pack()
    label_count_18.pack()
    label_count_19.pack()
    label_count_20.pack()
    label_count_25.pack()
    label_count_50.pack()

    label_count_1.pack_forget()
    label_count_2.pack_forget()
    label_count_3.pack_forget()
    label_count_4.pack_forget()
    label_count_5.pack_forget()
    label_count_6.pack_forget()
    label_count_7.pack_forget()
    label_count_8.pack_forget()
    label_count_9.pack_forget()
    label_count_10.pack_forget()
    label_count_11.pack_forget()
    label_count_12.pack_forget()
    label_count_13.pack_forget()
    label_count_14.pack_forget()
    label_count_15.pack_forget()
    label_count_16.pack_forget()
    label_count_17.pack_forget()
    label_count_18.pack_forget()
    label_count_19.pack_forget()
    label_count_20.pack_forget()
    label_count_25.pack_forget()
    label_count_50.pack_forget()

    reset3()


def reset3():
    """
    This function restores the original state
    :return:
    """

    label_1.pack()
    label_2.pack()
    label_3.pack()
    label_4.pack()
    label_5.pack()
    label_6.pack()
    label_7.pack()
    label_8.pack()
    label_9.pack()
    label_10.pack()
    label_11.pack()
    label_12.pack()
    label_13.pack()
    label_14.pack()
    label_15.pack()
    label_16.pack()
    label_17.pack()
    label_18.pack()
    label_19.pack()
    label_20.pack()
    label_25.pack()
    label_50.pack()

    label_1.pack_forget()
    label_2.pack_forget()
    label_3.pack_forget()
    label_4.pack_forget()
    label_5.pack_forget()
    label_6.pack_forget()
    label_7.pack_forget()
    label_8.pack_forget()
    label_9.pack_forget()
    label_10.pack_forget()
    label_11.pack_forget()
    label_12.pack_forget()
    label_13.pack_forget()
    label_14.pack_forget()
    label_15.pack_forget()
    label_16.pack_forget()
    label_17.pack_forget()
    label_18.pack_forget()
    label_19.pack_forget()
    label_20.pack_forget()
    label_25.pack_forget()
    label_50.pack_forget()

    reset4()


def reset4():
    """
    This function restores the original state
    :return
    """

    button_next.pack()
    button_plus.pack()
    button_minus.pack()

    button_next.pack_forget()
    button_plus.pack_forget()
    button_minus.pack_forget()

    label_welcome['text'] = "Please select a game mode"

    label_single.place(x=470, y=120, height=30, width=110)
    label_double.place(x=580, y=120, height=30, width=110)
    label_triple.place(x=690, y=120, height=30, width=110)
    switch_button.place(x=320, y=120, height=30, width=100)
    button_start.place(x=850, y=120, height=30, width=110)

    label_single['bg'] = "yellow"
    label_double['bg'] = "white"
    label_triple['bg'] = "white"


def switch_modes():
    """
    This function switches the different modes
    :return:
    """
    if label_single['bg'] == "yellow":
        label_single['bg'] = "white"
        label_double['bg'] = "yellow"

    elif label_double['bg'] == "yellow":
        label_double['bg'] = "white"
        label_triple['bg'] = "yellow"

    elif label_triple['bg'] == "yellow":
        label_triple['bg'] = "white"
        label_single['bg'] = "yellow"


def plus():
    """
    This function increments the labels
    :return:
    """
    # check each label
    if label_count_1['bg'] == "yellow":
        number = int(label_count_1['text'])
        if number < 3:
            number += 1
            label_count_1['text'] = number

    elif label_count_2['bg'] == "yellow":
        number = int(label_count_2['text'])
        if number < 3:
            number += 1
            label_count_2['text'] = number

    elif label_count_3['bg'] == "yellow":
        number = int(label_count_3['text'])
        if number < 3:
            number += 1
            label_count_3['text'] = number

    elif label_count_4['bg'] == "yellow":
        number = int(label_count_4['text'])
        if number < 3:
            number += 1
            label_count_4['text'] = number

    elif label_count_5['bg'] == "yellow":
        number = int(label_count_5['text'])
        if number < 3:
            number += 1
            label_count_5['text'] = number

    else:
        plus2()


def plus2():
    """
    This function increments the labels
    :return:
    """
    if label_count_6['bg'] == "yellow":
        number = int(label_count_6['text'])
        if number < 3:
            number += 1
            label_count_6['text'] = number

    elif label_count_7['bg'] == "yellow":
        number = int(label_count_7['text'])
        if number < 3:
            number += 1
            label_count_7['text'] = number

    elif label_count_8['bg'] == "yellow":
        number = int(label_count_8['text'])
        if number < 3:
            number += 1
            label_count_8['text'] = number

    elif label_count_9['bg'] == "yellow":
        number = int(label_count_9['text'])
        if number < 3:
            number += 1
            label_count_9['text'] = number

    elif label_count_10['bg'] == "yellow":
        number = int(label_count_10['text'])
        if number < 3:
            number += 1
            label_count_10['text'] = number

    else:
        plus3()


def plus3():
    """
    This function increments the labels
    :return:
    """
    if label_count_11['bg'] == "yellow":
        number = int(label_count_11['text'])
        if number < 3:
            number += 1
            label_count_11['text'] = number

    elif label_count_12['bg'] == "yellow":
        number = int(label_count_12['text'])
        if number < 3:
            number += 1
            label_count_12['text'] = number

    elif label_count_13['bg'] == "yellow":
        number = int(label_count_13['text'])
        if number < 3:
            number += 1
            label_count_13['text'] = number

    elif label_count_14['bg'] == "yellow":
        number = int(label_count_14['text'])
        if number < 3:
            number += 1
            label_count_14['text'] = number

    elif label_count_15['bg'] == "yellow":
        number = int(label_count_15['text'])
        if number < 3:
            number += 1
            label_count_15['text'] = number

    else:
        plus4()


def plus4():
    """
    This function increments the labels
    :return:
    """
    if label_count_16['bg'] == "yellow":
        number = int(label_count_16['text'])
        if number < 3:
            number += 1
            label_count_16['text'] = number

    elif label_count_17['bg'] == "yellow":
        number = int(label_count_17['text'])
        if number < 3:
            number += 1
            label_count_17['text'] = number

    elif label_count_18['bg'] == "yellow":
        number = int(label_count_18['text'])
        if number < 3:
            number += 1
            label_count_18['text'] = number

    elif label_count_19['bg'] == "yellow":
        number = int(label_count_19['text'])
        if number < 3:
            number += 1
            label_count_19['text'] = number

    elif label_count_20['bg'] == "yellow":
        number = int(label_count_20['text'])
        if number < 3:
            number += 1
            label_count_20['text'] = number

    else:
        plus5()


def plus5():
    """
    This function increments the labels
    :return:
    """
    if label_count_25['bg'] == "yellow":
        number = int(label_count_25['text'])
        if number < 3:
            number += 1
            label_count_25['text'] = number

    elif label_count_50['bg'] == "yellow":
        number = int(label_count_50['text'])
        if number < 3:
            number += 1
            label_count_50['text'] = number


def minus():
    """
    This function decrements the labels
    :return:
    """
    if label_count_1['bg'] == "yellow":
        number = int(label_count_1['text'])
        if number > 0:
            number -= 1
            label_count_1['text'] = number

    elif label_count_2['bg'] == "yellow":
        number = int(label_count_2['text'])
        if number > 0:
            number -= 1
            label_count_2['text'] = number

    elif label_count_3['bg'] == "yellow":
        number = int(label_count_3['text'])
        if number > 0:
            number -= 1
            label_count_3['text'] = number

    elif label_count_4['bg'] == "yellow":
        number = int(label_count_4['text'])
        if number > 0:
            number -= 1
            label_count_4['text'] = number

    elif label_count_5['bg'] == "yellow":
        number = int(label_count_5['text'])
        if number > 0:
            number -= 1
            label_count_5['text'] = number

    else:
        minus2()


def minus2():
    """
    This function decrements the labels
    :return:
    """
    if label_count_6['bg'] == "yellow":
        number = int(label_count_6['text'])
        if number > 0:
            number -= 1
            label_count_6['text'] = number

    elif label_count_7['bg'] == "yellow":
        number = int(label_count_7['text'])
        if number > 0:
            number -= 1
            label_count_7['text'] = number

    elif label_count_8['bg'] == "yellow":
        number = int(label_count_8['text'])
        if number > 0:
            number -= 1
            label_count_8['text'] = number

    elif label_count_9['bg'] == "yellow":
        number = int(label_count_9['text'])
        if number > 0:
            number -= 1
            label_count_9['text'] = number

    elif label_count_10['bg'] == "yellow":
        number = int(label_count_10['text'])
        if number > 0:
            number -= 1
            label_count_10['text'] = number

    else:
        minus3()


def minus3():
    """
    This function decrements the labels
    :return:
    """

    if label_count_11['bg'] == "yellow":
        number = int(label_count_11['text'])
        if number > 0:
            number -= 1
            label_count_11['text'] = number

    elif label_count_12['bg'] == "yellow":
        number = int(label_count_12['text'])
        if number > 0:
            number -= 1
            label_count_12['text'] = number

    elif label_count_13['bg'] == "yellow":
        number = int(label_count_13['text'])
        if number > 0:
            number -= 1
            label_count_13['text'] = number

    elif label_count_14['bg'] == "yellow":
        number = int(label_count_14['text'])
        if number > 0:
            number -= 1
            label_count_14['text'] = number

    elif label_count_15['bg'] == "yellow":
        number = int(label_count_15['text'])
        if number > 0:
            number -= 1
            label_count_15['text'] = number

    else:
        minus4()


def minus4():
    """
    This function decrements the labels
    :return:
    """
    if label_count_16['bg'] == "yellow":
        number = int(label_count_16['text'])
        if number > 0:
            number -= 1
            label_count_16['text'] = number

    elif label_count_17['bg'] == "yellow":
        number = int(label_count_17['text'])
        if number > 0:
            number -= 1
            label_count_17['text'] = number

    elif label_count_18['bg'] == "yellow":
        number = int(label_count_18['text'])
        if number > 0:
            number -= 1
            label_count_18['text'] = number

    elif label_count_19['bg'] == "yellow":
        number = int(label_count_19['text'])
        if number > 0:
            number -= 1
            label_count_19['text'] = number

    elif label_count_20['bg'] == "yellow":
        number = int(label_count_20['text'])
        if number > 0:
            number -= 1
            label_count_20['text'] = number

    else:
        minus5()


def minus5():
    """
    This function decrements the labels
    :return:
    """
    if label_count_25['bg'] == "yellow":
        number = int(label_count_25['text'])
        if number > 0:
            number -= 1
            label_count_25['text'] = number

    elif label_count_50['bg'] == "yellow":
        number = int(label_count_50['text'])
        if number > 0:
            number -= 1
            label_count_50['text'] = number


def next_label1():
    """
    This function moves on to the next field
    :return:
    """
    if label_count_1['bg'] == "yellow":
        label_count_1['bg'] = "white"
        label_count_2['bg'] = "yellow"

    elif label_count_2['bg'] == "yellow":
        label_count_2['bg'] = "white"
        label_count_3['bg'] = "yellow"

    elif label_count_3['bg'] == "yellow":
        label_count_3['bg'] = "white"
        label_count_4['bg'] = "yellow"

    elif label_count_4['bg'] == "yellow":
        label_count_4['bg'] = "white"
        label_count_5['bg'] = "yellow"

    elif label_count_5['bg'] == "yellow":
        label_count_5['bg'] = "white"
        label_count_6['bg'] = "yellow"

    elif label_count_6['bg'] == "yellow":
        label_count_6['bg'] = "white"
        label_count_7['bg'] = "yellow"

    elif label_count_7['bg'] == "yellow":
        label_count_7['bg'] = "white"
        label_count_8['bg'] = "yellow"

    elif label_count_8['bg'] == "yellow":
        label_count_8['bg'] = "white"
        label_count_9['bg'] = "yellow"

    elif label_count_9['bg'] == "yellow":
        label_count_9['bg'] = "white"
        label_count_10['bg'] = "yellow"

    elif label_count_10['bg'] == "yellow":
        label_count_10['bg'] = "white"
        label_count_11['bg'] = "yellow"

    elif label_count_11['bg'] == "yellow":
        label_count_11['bg'] = "white"
        label_count_12['bg'] = "yellow"

    else:
        next_label2()


def next_label2():
    """
    This function moves on to the next field
    :return:
    """
    if label_count_12['bg'] == "yellow":
        label_count_12['bg'] = "white"
        label_count_13['bg'] = "yellow"

    elif label_count_13['bg'] == "yellow":
        label_count_13['bg'] = "white"
        label_count_14['bg'] = "yellow"

    elif label_count_14['bg'] == "yellow":
        label_count_14['bg'] = "white"
        label_count_15['bg'] = "yellow"

    elif label_count_15['bg'] == "yellow":
        label_count_15['bg'] = "white"
        label_count_16['bg'] = "yellow"

    elif label_count_16['bg'] == "yellow":
        label_count_16['bg'] = "white"
        label_count_17['bg'] = "yellow"

    elif label_count_17['bg'] == "yellow":
        label_count_17['bg'] = "white"
        label_count_18['bg'] = "yellow"

    elif label_count_18['bg'] == "yellow":
        label_count_18['bg'] = "white"
        label_count_19['bg'] = "yellow"

    else:
        next_label3()


def next_label3():
    """
    This function moves on to the next field
    :return:
    """
    if label_count_19['bg'] == "yellow":
        label_count_19['bg'] = "white"
        label_count_20['bg'] = "yellow"

    elif label_count_20['bg'] == "yellow":
        label_count_20['bg'] = "white"
        if label_single['bg'] == "yellow":
            label_count_25['bg'] = "yellow"
        elif label_double['bg'] == "yellow":
            label_count_50['bg'] = "yellow"
        else:
            end_game()

    elif label_count_25['bg'] == "yellow":
        end_game()

    elif label_count_50['bg'] == "yellow":
        end_game()


def end_game():
    """
    This function ends the game
    :return:
    """
    messagebox.showinfo("Info", "Around-the-clock finished.")

    create_excel()
    reset()


def start_game():
    """
    After the mode was chosen, the game can be started
    :return:
    """
    if label_single['bg'] == "yellow":
        label_selected_mode['text'] = "Selected mode: single"

    elif label_double['bg'] == "yellow":
        label_selected_mode['text'] = "Selected mode: double"

    elif label_triple['bg'] == "yellow":
        label_selected_mode['text'] = "Selected mode: triple"

    label_selected_mode.place(x=250, y=10, height=30, width=200)

    label_single.pack()
    label_double.pack()
    label_triple.pack()
    label_welcome.pack()
    button_start.pack()
    switch_button.pack()

    label_single.pack_forget()
    label_double.pack_forget()
    label_triple.pack_forget()
    label_welcome.pack_forget()
    button_start.pack_forget()
    switch_button.pack_forget()

    # enable player name
    label_player_1_name.place(x=10, y=10, height=30, width=150)

    # enable all buttons and labels for the game
    label_1.place(x=0, y=160, height=30, width=100)
    label_2.place(x=0, y=200, height=30, width=100)
    label_3.place(x=0, y=240, height=30, width=100)
    label_4.place(x=0, y=280, height=30, width=100)
    label_5.place(x=0, y=320, height=30, width=100)
    label_6.place(x=0, y=360, height=30, width=100)
    label_7.place(x=0, y=400, height=30, width=100)
    label_8.place(x=0, y=440, height=30, width=100)
    label_9.place(x=0, y=480, height=30, width=100)
    label_10.place(x=0, y=520, height=30, width=100)
    label_11.place(x=200, y=160, height=30, width=100)
    label_12.place(x=200, y=200, height=30, width=100)
    label_13.place(x=200, y=240, height=30, width=100)
    label_14.place(x=200, y=280, height=30, width=100)
    label_15.place(x=200, y=320, height=30, width=100)
    label_16.place(x=200, y=360, height=30, width=100)
    label_17.place(x=200, y=400, height=30, width=100)
    label_18.place(x=200, y=440, height=30, width=100)
    label_19.place(x=200, y=480, height=30, width=100)
    label_20.place(x=200, y=520, height=30, width=100)

    start_game2()


def start_game2():
    """
    After the mode was chosen, the game can be started
    :return:
    """
    label_count_1.place(x=110, y=160, height=30, width=30)
    label_count_2.place(x=110, y=200, height=30, width=30)
    label_count_3.place(x=110, y=240, height=30, width=30)
    label_count_4.place(x=110, y=280, height=30, width=30)
    label_count_5.place(x=110, y=320, height=30, width=30)
    label_count_6.place(x=110, y=360, height=30, width=30)
    label_count_7.place(x=110, y=400, height=30, width=30)
    label_count_8.place(x=110, y=440, height=30, width=30)
    label_count_9.place(x=110, y=480, height=30, width=30)
    label_count_10.place(x=110, y=520, height=30, width=30)
    label_count_11.place(x=310, y=160, height=30, width=30)
    label_count_12.place(x=310, y=200, height=30, width=30)
    label_count_13.place(x=310, y=240, height=30, width=30)
    label_count_14.place(x=310, y=280, height=30, width=30)
    label_count_15.place(x=310, y=320, height=30, width=30)
    label_count_16.place(x=310, y=360, height=30, width=30)
    label_count_17.place(x=310, y=400, height=30, width=30)
    label_count_18.place(x=310, y=440, height=30, width=30)
    label_count_19.place(x=310, y=480, height=30, width=30)
    label_count_20.place(x=310, y=520, height=30, width=30)

    button_minus.place(x=800, y=240, height=60, width=60)
    button_plus.place(x=880, y=240, height=60, width=60)
    button_next.place(x=800, y=160, height=60, width=140)

    if label_single['bg'] == "yellow":
        label_count_25.place(x=510, y=160, height=30, width=30)
        label_25.place(x=400, y=160, height=30, width=100)

    elif label_double['bg'] == "yellow":
        label_count_50.place(x=510, y=160, height=30, width=30)
        label_50.place(x=400, y=160, height=30, width=100)


def enter_name():
    """

    :return:
    """
    name = input_name1.get()
    if name != "":
        input_name1.pack()
        input_name1.pack_forget()
        button_entering_name.pack()
        button_entering_name.pack_forget()
        label_welcome['text'] = "Please select a game mode"

        label_single.place(x=470, y=120, height=30, width=110)
        label_double.place(x=580, y=120, height=30, width=110)
        label_triple.place(x=690, y=120, height=30, width=110)
        switch_button.place(x=320, y=120, height=30, width=100)
        button_start.place(x=850, y=120, height=30, width=110)

        label_player_1_name['text'] = name

    else:
        messagebox.showinfo("Error", "Please enter your name.")


if __name__ == "__main__":
    # configure the window to generate
    gui = Tk()
    gui.geometry('1275x645')
    gui.resizable(width=False, height=False)
    gui.title("Around the clock")
    gui.configure(background='grey')

    # define the exit - button
    exit_button = Button(gui, text="End Game", command=button_exit, fg="black", bg="lightgreen",
                         font=('Arial', 10, 'bold'))
    exit_button.place(x=1175, y=0, height=80, width=100)

    # welcome label
    label_welcome = Label(gui, text="Welcome to the Around-the-clock game!"
                                    " Please enter your name.", bg="grey",
                          font=('Arial', 14))
    label_welcome.place(x=337.5, y=20, height=30, width=600)

    button_entering_name = Button(gui, text="Continue", bg="yellow",
                                  font=('Arial', 10), command=enter_name)
    button_entering_name.place(x=737.5, y=140, height=30, width=100)
    # label and button for entering name

    input_name1 = Entry(gui, bd=1, font=('Arial', 13))

    input_name1.place(x=537.5, y=140, height=30, width=200)
    label_player_1_name = Label(gui, text="", bg="yellow",
                                fg="black", font=('Arial', 13, 'bold'))

    # Label for Single, Double and Single, also button to switch
    switch_button = Button(gui, text="Switch", command=switch_modes, fg="black", bg="lightblue",
                           font=('Arial', 13, 'bold'))
    label_single = Label(gui, text="Single", bg="yellow", fg="black", font=('Arial', 13, 'bold'))
    label_double = Label(gui, text="Double", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_triple = Label(gui, text="Triple", bg="white", fg="black", font=('Arial', 13, 'bold'))

    # define start button
    button_start = Button(gui, text="Start", bg="lightgreen", fg="black",
                          font=('Arial', 13, 'bold'), command=start_game)

    # reset - button
    reset_button = Button(gui, text="Reset", bd=4, fg="black", bg="red", font=('Arial', 11),
                          command=reset)

    reset_button.place(x=1175, y=110, height=30, width=100)

    # define label for selected mode
    label_selected_mode = Label(gui, text="", font=('Arial', 13))

    # labels for all fields
    label_20 = Label(gui, text="20", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_19 = Label(gui, text="19", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_18 = Label(gui, text="18", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_17 = Label(gui, text="17", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_16 = Label(gui, text="16", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_15 = Label(gui, text="15", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_14 = Label(gui, text="14", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_13 = Label(gui, text="13", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_12 = Label(gui, text="12", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_11 = Label(gui, text="11", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_10 = Label(gui, text="10", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_9 = Label(gui, text="9", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_8 = Label(gui, text="8", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_7 = Label(gui, text="7", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_6 = Label(gui, text="6", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_5 = Label(gui, text="5", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_4 = Label(gui, text="4", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_3 = Label(gui, text="3", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_2 = Label(gui, text="2", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_1 = Label(gui, text="1", bg="green", fg="black", font=('Arial', 13, 'bold'))
    label_50 = Label(gui, text="BULL", bg="red", fg="black", font=('Arial', 13, 'bold'))
    label_25 = Label(gui, text="25", bg="green", fg="black", font=('Arial', 13, 'bold'))

    # labels for results of the thrown darts
    label_count_20 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_19 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_18 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_17 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_16 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_15 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_14 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_13 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_12 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_11 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_10 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_9 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_8 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_7 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_6 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_5 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_4 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_3 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_2 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_1 = Label(gui, text="0", bg="yellow", fg="black", font=('Arial', 13, 'bold'))
    label_count_50 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))
    label_count_25 = Label(gui, text="0", bg="white", fg="black", font=('Arial', 13, 'bold'))

    # buttons for plus and minus and next
    button_plus = Button(gui, text="+", bd=4, fg="black", bg="lightgreen", font=('Arial', 11),
                         command=plus)
    button_minus = Button(gui, text="-", bd=4, fg="black", bg="red", font=('Arial', 11),
                          command=minus)

    button_next = Button(gui, text="Next", bd=4, fg="white", bg="black", font=('Arial', 11),
                         command=next_label1)

    gui.mainloop()
