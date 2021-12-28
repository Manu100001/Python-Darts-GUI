#!/usr/bin/python3
"""
This script will help you play darts.
If you don't have an electric dartboard but a normal one,
this script will help you to calculate the scores.

:author: Manuel Milde manuelmilde@gmx.net
:copyright: 2021 Manuel Milde
"""
from tkinter import *
from tkinter import messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# color for excel
greenFill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')

# months for creating timebased excel
months = ['Januar', 'Februar', 'März', 'April', 'Mai', 'Juni', 'Juli', 'August', 'September', 'Oktober', 'November',
          'Dezember']

# safe darts and score for each player
player1 = []
player2 = []
player3 = []
player4 = []

# index for players for excel sheet
start_index_1 = 0
start_index_2 = 0
start_index_3 = 0
start_index_4 = 0

# dictionaries for each player
player1_kpis = [{"Score": 0, "Darts": 0, "180": 0, "140": 0, "100": 0, "80": 0, "60": 0}]
player2_kpis = [{"Score": 0, "Darts": 0, "180": 0, "140": 0, "100": 0, "80": 0, "60": 0}]
player3_kpis = [{"Score": 0, "Darts": 0, "180": 0, "140": 0, "100": 0, "80": 0, "60": 0}]
player4_kpis = [{"Score": 0, "Darts": 0, "180": 0, "140": 0, "100": 0, "80": 0, "60": 0}]


def create_excel():
    """

    :return:
    """
    if not os.path.isdir("Spielstände"):
        os.mkdir("Spielstände")

    if not os.path.isdir("Spielstände/Scoring"):
        os.mkdir("Spielstände/Scoring")

    # current year
    current_year = datetime.now().strftime('%Y')
    if not os.path.isdir("Spielstände/Scoring/" + current_year):
        os.mkdir("Spielstände/Scoring/" + current_year)

    # current month
    current_month = datetime.now().strftime('%m')
    month_name = months[int(current_month) - 1]
    if not os.path.isdir("Spielstände/Scoring/" + current_year + "/" + month_name):
        os.mkdir("Spielstände/Scoring/" + current_year + "/" + month_name)

    # current day
    current_day = int(datetime.now().strftime('%d'))
    date = str(current_day) + "." + str(current_month)
    if not os.path.isdir("Spielstände/Scoring/" + current_year + "/" + month_name + "/" + date):
        os.mkdir("Spielstände/Scoring/" + current_year + "/" + month_name + "/" + date)

    # create new score - file
    time = datetime.now().strftime('%H-%M-%S')

    excel_file = Workbook()
    sheet = excel_file.create_sheet('Scoring')
    path = "Spielstände/Scoring/" + current_year + "/" + month_name + "/" + date + "/" + time + ".xlsx"
    # set standards
    sheet.cell(row=3, column=2).value = "Spieler"
    sheet.cell(row=3, column=3).value = "Average"
    sheet.cell(row=3, column=4).value = "180"
    sheet.cell(row=3, column=5).value = "140+"
    sheet.cell(row=3, column=6).value = "100+"
    sheet.cell(row=3, column=7).value = "80+"
    sheet.cell(row=3, column=8).value = "60+"

    sheet.cell(row=3, column=9).value = "Geworfene Punkte"
    sheet.cell(row=3, column=10).value = "Geworfene Darts"

    sheet['B3'].fill = greenFill
    sheet['C3'].fill = greenFill
    sheet['D3'].fill = greenFill
    sheet['E3'].fill = greenFill
    sheet['F3'].fill = greenFill
    sheet['G3'].fill = greenFill
    sheet['H3'].fill = greenFill
    sheet['I3'].fill = greenFill
    sheet['J3'].fill = greenFill

    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 18

    # logic for excel - file
    sheet.cell(row=4, column=2).value = label_player_1_name['text']
    sheet.cell(row=5, column=2).value = label_player_2_name['text']
    sheet.cell(row=6, column=2).value = label_player_3_name['text']
    sheet.cell(row=7, column=2).value = label_player_4_name['text']

    # show data in excel
    # player 1
    sheet.cell(row=4, column=3).value = round((player1_kpis[0]['Score'] / player1_kpis[0]['Darts']) * 3, 2)
    sheet.cell(row=4, column=4).value = player1_kpis[0]['180']
    sheet.cell(row=4, column=5).value = player1_kpis[0]['140']
    sheet.cell(row=4, column=6).value = player1_kpis[0]['100']
    sheet.cell(row=4, column=7).value = player1_kpis[0]['80']
    sheet.cell(row=4, column=8).value = player1_kpis[0]['60']

    sheet.cell(row=4, column=9).value = player1_kpis[0]['Score']
    sheet.cell(row=4, column=10).value = player1_kpis[0]['Darts']

    # player 2
    sheet.cell(row=5, column=3).value = round((player2_kpis[0]['Score'] / player2_kpis[0]['Darts']) * 3, 2)
    sheet.cell(row=5, column=4).value = player2_kpis[0]['180']
    sheet.cell(row=5, column=5).value = player2_kpis[0]['140']
    sheet.cell(row=5, column=6).value = player2_kpis[0]['100']
    sheet.cell(row=5, column=7).value = player2_kpis[0]['80']
    sheet.cell(row=5, column=8).value = player2_kpis[0]['60']

    sheet.cell(row=5, column=9).value = player2_kpis[0]['Score']
    sheet.cell(row=5, column=10).value = player2_kpis[0]['Darts']

    # player 3
    if label_player_3_name['text'] != "":
        sheet.cell(row=6, column=3).value = round((player3_kpis[0]['Score'] / player3_kpis[0]['Darts']) * 3, 2)
        sheet.cell(row=6, column=4).value = player3_kpis[0]['180']
        sheet.cell(row=6, column=5).value = player3_kpis[0]['140']
        sheet.cell(row=6, column=6).value = player3_kpis[0]['100']
        sheet.cell(row=6, column=7).value = player3_kpis[0]['80']
        sheet.cell(row=6, column=8).value = player3_kpis[0]['60']

        sheet.cell(row=6, column=9).value = player3_kpis[0]['Score']
        sheet.cell(row=6, column=10).value = player3_kpis[0]['Darts']

    # player 4
    if label_player_4_name['text'] != "":
        sheet.cell(row=7, column=3).value = round((player4_kpis[0]['Score'] / player4_kpis[0]['Darts']) * 3, 2)
        sheet.cell(row=7, column=4).value = player4_kpis[0]['180']
        sheet.cell(row=7, column=5).value = player4_kpis[0]['140']
        sheet.cell(row=7, column=6).value = player4_kpis[0]['100']
        sheet.cell(row=7, column=7).value = player4_kpis[0]['80']
        sheet.cell(row=7, column=8).value = player4_kpis[0]['60']

        sheet.cell(row=7, column=9).value = player4_kpis[0]['Score']
        sheet.cell(row=7, column=10).value = player4_kpis[0]['Darts']

    # save excel - file
    excel_file.save(path)


def save_score():
    """

    :return:
    """
    # update data
    for item in player1:
        score = item['Score']
        darts = item['Darts']

        player1_kpis[0]['Score'] += score
        player1_kpis[0]['Darts'] += darts

        if score >= 60:
            if score >= 80:
                if score >= 100:
                    if score >= 140:
                        if score == 180:
                            player1_kpis[0]['180'] += 1
                        else:
                            player1_kpis[0]['140'] += 1
                    else:
                        player1_kpis[0]['100'] += 1
                else:
                    player1_kpis[0]['80'] += 1
            else:
                player1_kpis[0]['60'] += 1

    # update player 2
    for item in player2:
        score = item['Score']
        darts = item['Darts']

        player2_kpis[0]['Score'] += score
        player2_kpis[0]['Darts'] += darts

        if score >= 60:
            if score >= 80:
                if score >= 100:
                    if score >= 140:
                        if score == 180:
                            player2_kpis[0]['180'] += 1
                        else:
                            player2_kpis[0]['140'] += 1
                    else:
                        player2_kpis[0]['100'] += 1
                else:
                    player2_kpis[0]['80'] += 1
            else:
                player2_kpis[0]['60'] += 1

    # update player 3
    for item in player3:
        score = item['Score']
        darts = item['Darts']

        player3_kpis[0]['Score'] += score
        player3_kpis[0]['Darts'] += darts

        if score >= 60:
            if score >= 80:
                if score >= 100:
                    if score >= 140:
                        if score == 180:
                            player3_kpis[0]['180'] += 1
                        else:
                            player3_kpis[0]['140'] += 1
                    else:
                        player3_kpis[0]['100'] += 1
                else:
                    player3_kpis[0]['80'] += 1
            else:
                player3_kpis[0]['60'] += 1

    # update player 4
    for item in player4:
        score = item['Score']
        darts = item['Darts']

        player4_kpis[0]['Score'] += score
        player4_kpis[0]['Darts'] += darts

        if score >= 60:
            if score >= 80:
                if score >= 100:
                    if score >= 140:
                        if score == 180:
                            player4_kpis[0]['180'] += 1
                        else:
                            player4_kpis[0]['140'] += 1
                    else:
                        player4_kpis[0]['100'] += 1
                else:
                    player4_kpis[0]['80'] += 1
            else:
                player4_kpis[0]['60'] += 1


def add_player1(result, dart):
    """

    :return:
    """
    player1.append({"Score": result, "Darts": dart})
    return


def add_player2(result, dart):
    """

    :return:
    """
    player2.append({"Score": result, "Darts": dart})
    return


def add_player3(result, dart):
    """

    :return:
    """
    player3.append({"Score": result, "Darts": dart})
    return


def add_player4(result, dart):
    """

    :return:
    """
    player4.append({"Score": result, "Darts": dart})
    return


def T20():
    """

    :return:
    """
    label_dart_score['text'] = "60"


def D20():
    """

    :return:
    """
    label_dart_score['text'] = "40"


def S20():
    """

    :return:
    """
    label_dart_score['text'] = "20"


def T19():
    """"
    """
    label_dart_score['text'] = "57"


def D19():
    """

    :return:
    """
    label_dart_score['text'] = "38"


def S19():
    """

    :return:
    """
    label_dart_score['text'] = "19"


def T18():
    """

    :return:
    """
    label_dart_score['text'] = "54"


def D18():
    """

    :return:
    """
    label_dart_score['text'] = "36"


def S18():
    """

    :return:
    """
    label_dart_score['text'] = "18"


def T17():
    """

    :return:
    """
    label_dart_score['text'] = "51"


def D17():
    """

    :return:
    """
    label_dart_score['text'] = "34"


def S17():
    """

    :return:
    """
    label_dart_score['text'] = "17"


def T16():
    """

    :return:
    """
    label_dart_score['text'] = "48"


def D16():
    """

    :return:
    """
    label_dart_score['text'] = "32"


def S16():
    """

    :return:
    """
    label_dart_score['text'] = "16"


def T15():
    """

    :return:
    """
    label_dart_score['text'] = "45"


def D15():
    """

    :return:
    """
    label_dart_score['text'] = "30"


def S15():
    """

    :return:
    """
    label_dart_score['text'] = "15"


def T14():
    """

    :return:
    """
    label_dart_score['text'] = "42"


def D14():
    """

    :return:
    """
    label_dart_score['text'] = "28"


def S14():
    """

    :return:
    """
    label_dart_score['text'] = "14"


def T13():
    """

    :return:
    """
    label_dart_score['text'] = "39"


def D13():
    """

    :return:
    """
    label_dart_score['text'] = "26"


def S13():
    """

    :return:
    """
    label_dart_score['text'] = "13"


def T12():
    """

    :return:
    """
    label_dart_score['text'] = "36"


def D12():
    """

    :return:
    """
    label_dart_score['text'] = "24"


def S12():
    """

    :return:
    """
    label_dart_score['text'] = "12"


def T11():
    """

    :return:
    """
    label_dart_score['text'] = "33"


def D11():
    """

    :return:
    """
    label_dart_score['text'] = "22"


def S11():
    """

    :return:
    """
    label_dart_score['text'] = "11"


def T10():
    """

    :return:
    """
    label_dart_score['text'] = "30"


def D10():
    """

    :return:
    """
    label_dart_score['text'] = "20"


def S10():
    """

    :return:
    """
    label_dart_score['text'] = "10"


def T9():
    """

    :return:
    """
    label_dart_score['text'] = "27"


def D9():
    """

    :return:
    """
    label_dart_score['text'] = "18"


def S9():
    """

    :return:
    """
    label_dart_score['text'] = "9"


def T8():
    """

    :return:
    """
    label_dart_score['text'] = "24"


def D8():
    """

    :return:
    """
    label_dart_score['text'] = "16"


def S8():
    """

    :return:
    """
    label_dart_score['text'] = "8"


def T7():
    """

    :return:
    """
    label_dart_score['text'] = "21"


def D7():
    """

    :return:
    """
    label_dart_score['text'] = "14"


def S7():
    """

    :return:
    """
    label_dart_score['text'] = "7"


def T6():
    """

    :return:
    """
    label_dart_score['text'] = "18"


def D6():
    """

    :return:
    """
    label_dart_score['text'] = "12"


def S6():
    """

    :return:
    """
    label_dart_score['text'] = "6"


def T5():
    """

    :return:
    """
    label_dart_score['text'] = "15"


def D5():
    """

    :return:
    """
    label_dart_score['text'] = "10"


def S5():
    """

    :return:
    """
    label_dart_score['text'] = "5"


def T4():
    """

    :return:
    """
    label_dart_score['text'] = "12"


def D4():
    """

    :return:
    """
    label_dart_score['text'] = "8"


def S4():
    """

    :return:
    """
    label_dart_score['text'] = "4"


def T3():
    """

    :return:
    """
    label_dart_score['text'] = "9"


def D3():
    """

    :return:
    """
    label_dart_score['text'] = "6"


def S3():
    """

    :return:
    """
    label_dart_score['text'] = "3"


def T2():
    """

    :return:
    """
    label_dart_score['text'] = "6"


def D2():
    """

    :return:
    """
    label_dart_score['text'] = "4"


def S2():
    """

    :return:
    """
    label_dart_score['text'] = "2"


def T1():
    """

    :return:
    """
    label_dart_score['text'] = "3"


def D1():
    """

    :return:
    """
    label_dart_score['text'] = "2"


def S1():
    """

    :return:
    """
    label_dart_score['text'] = "1"


def Bull():
    """

    :return:
    """
    label_dart_score['text'] = "50"


def Single_Bull():
    """

    :return:
    """
    label_dart_score['text'] = "25"


def null():
    """

    :return:
    """
    label_dart_score['text'] = "0"


def button_exit():
    """
    This function creates an exit - button for the gui
    :return:
    """
    if not any(isinstance(window, Toplevel) for window in gui.winfo_children()):
        exit_window = Toplevel(gui)
        exit_window.geometry('250x150')
        exit_window.resizable(width=0, height=0)
        exit_window.title("Beenden?")

        label_exit = Label(exit_window, text="Spiel beenden?", font=('Arial', 11))
        button_ja = Button(exit_window, text="Ja", command=exit_window.quit, font=('Arial', 10, 'bold'), bg="white",
                           fg="green")
        button_nein = Button(exit_window, text="Nein", command=exit_window.destroy, font=('Arial', 10, 'bold'),
                             bg="white", fg="red")

        label_exit.place(x=80, y=0, width=100, height=50)
        button_ja.place(x=50, y=60, width=50, height=50)
        button_nein.place(x=150, y=60, width=50, height=50)

    else:
        messagebox.showinfo("Info", "You already clicked on \"Beenden\"!")


def button_name_1():
    """

    :return:
    """
    name = eingabefeld_p1.get()
    eingabefeld_p1.delete("0", "end")
    label_player_1_name['text'] = name


def button_name_2():
    """

    :return:
    """
    name = eingabefeld_p2.get()
    eingabefeld_p2.delete("0", "end")
    label_player_2_name['text'] = name


def button_name_3():
    """


    :return:
    """
    name = eingabefeld_p3.get()
    eingabefeld_p3.delete("0", "end")
    label_player_3_name['text'] = name


def button_name_4():
    """

    :return:
    """
    name = eingabefeld_p4.get()
    eingabefeld_p4.delete("0", "end")
    label_player_4_name['text'] = name


def button_switch_score_inc():
    """

    :return:
    """
    label_switch_score['text'] = "501"
    label_1_score['text'] = "501"
    label_2_score['text'] = "501"
    label_3_score['text'] = "501"
    label_4_score['text'] = "501"


def button_switch_score_dec():
    """

    :return:
    """
    label_switch_score['text'] = "301"
    label_1_score['text'] = "301"
    label_2_score['text'] = "301"
    label_3_score['text'] = "301"
    label_4_score['text'] = "301"


def button_stop_game_function():
    """

    :return:
    """
    button_stop_game.pack()
    button_stop_game.pack_forget()

    next_button.pack()
    next_button.pack_forget()

    button_start_game.pack()
    button_start_game.place(x=850, y=90, height=30, width=100)

    button_name_1.pack()
    button_name_2.pack()
    button_name_3.pack()
    button_name_4.pack()

    button_name_1.place(x=850, y=5, height=30)
    button_name_2.place(x=850, y=35, height=30)
    button_name_3.place(x=1000, y=5, height=30)
    button_name_4.place(x=1000, y=35, height=30)

    eingabefeld_p1.pack()
    eingabefeld_p2.pack()
    eingabefeld_p3.pack()
    eingabefeld_p4.pack()

    eingabefeld_p1.place(x=750, y=5, width=100, height=30)
    eingabefeld_p2.place(x=750, y=35, width=100, height=30)
    eingabefeld_p3.place(x=900, y=5, width=100, height=30)
    eingabefeld_p4.place(x=900, y=35, width=100, height=30)

    label_switch_score.pack()
    button_switch_score_dec.pack()
    button_switch_score_inc.pack()

    label_switch_score.place(x=1030, y=90, height=30, width=110)
    button_switch_score_dec.place(x=1140, y=90, height=30, width=30)
    button_switch_score_inc.place(x=1170, y=90, height=30, width=30)

    label_dart_score.pack()
    button_dart_score.pack()

    label_dart_score.pack_forget()
    button_dart_score.pack_forget()

    label_first_dart.pack()
    label_second_dart.pack()
    label_third_dart.pack()
    zwischen_label.pack()

    label_first_dart.pack_forget()
    label_second_dart.pack_forget()
    label_third_dart.pack_forget()
    zwischen_label.pack_forget()


def button_start_game_function():
    """

    :return:
    """
    label_1_score['bg'] = "yellow"
    label_2_score['bg'] = "white"
    label_3_score['bg'] = "white"
    label_4_score['bg'] = "white"

    for item in player1:
        player1.remove(item)
    for item in player2:
        player2.remove(item)
    for item in player3:
        player3.remove(item)
    for item in player4:
        player4.remove(item)

    if int(label_switch_score['text']) == 501:
        label_1_score['text'] = "501"
        label_2_score['text'] = "501"
    else:
        label_1_score['text'] = "301"
        label_2_score['text'] = "301"

    if label_player_3_name['text'] == "Player 3: " or label_player_3_name['text'] == "":
        label_player_3_name['text'] = ""
        label_3_score['text'] = ""
    else:
        if int(label_switch_score['text']) == 501:
            label_3_score['text'] = "501"
        else:
            label_3_score['text'] = "301"

    if label_player_4_name['text'] == "Player 4: " or label_player_4_name['text'] == "":
        label_player_4_name['text'] = ""
        label_4_score['text'] = ""
    else:
        if int(label_switch_score['text']) == 501:
            label_4_score['text'] = "501"
        else:
            label_4_score['text'] = "301"

    button_start_game.pack()
    button_start_game.pack_forget()
    button_name_1.pack()
    button_name_1.pack_forget()
    button_name_2.pack()
    button_name_2.pack_forget()
    button_name_3.pack()
    button_name_3.pack_forget()
    button_name_4.pack()
    button_name_4.pack_forget()

    eingabefeld_p1.pack()
    eingabefeld_p1.pack_forget()
    eingabefeld_p2.pack()
    eingabefeld_p2.pack_forget()
    eingabefeld_p3.pack()
    eingabefeld_p3.pack_forget()
    eingabefeld_p4.pack()
    eingabefeld_p4.pack_forget()

    button_switch_score_inc.pack()
    button_switch_score_inc.pack_forget()
    button_switch_score_dec.pack()
    button_switch_score_dec.pack_forget()
    label_switch_score.pack()
    label_switch_score.pack_forget()

    button_stop_game.pack()
    button_stop_game.place(x=950, y=90, height=30, width=100)

    next_button.pack()
    next_button.place(x=610, y=60, height=30, width=100)

    label_dart_score.pack()
    button_dart_score.pack()

    label_dart_score.place(x=0, y=300, height=30, width=90)
    button_dart_score.place(x=90, y=300, height=30, width=80)

    label_first_dart.pack()
    label_second_dart.pack()
    label_third_dart.pack()
    zwischen_label.pack()

    label_first_dart.place(x=210, y=300, height=30, width=30)
    label_second_dart.place(x=250, y=300, height=30, width=30)
    label_third_dart.place(x=290, y=300, height=30, width=30)
    zwischen_label.place(x=340, y=300, height=30, width=100)


def next_button():
    """

    :return:
    """
    zwischen_label['text'] = "0"
    label_first_dart['bg'] = "yellow"
    label_second_dart['bg'] = "white"
    label_third_dart['bg'] = "white"

    # check label 1
    if label_1_score['bg'] == "yellow":
        label_1_score['bg'] = "white"

        if int(label_2_score['text']) == 0:
            # is label 3 set?
            if label_3_score['text'] != "":
                if label_4_score['text'] != "":
                    if int(label_3_score['text']) == 0:
                        if int(label_4_score['text']) == 0:
                            label_1_score['bg'] = "yellow"
                        else:
                            label_4_score['bg'] = "yellow"
                    else:
                        label_3_score['bg'] = "yellow"
        else:
            label_2_score['bg'] = "yellow"
        return

    # check label 2
    if label_2_score['bg'] == "yellow":
        label_2_score['bg'] = "white"

        if label_3_score['text'] != "":
            if label_4_score['text'] != "":
                if int(label_3_score['text']) == 0:
                    if int(label_4_score['text']) == 0:
                        label_1_score['bg'] = "yellow"
                    else:
                        label_4_score['bg'] = "yellow"
                else:
                    label_3_score['bg'] = "yellow"
            elif int(label_3_score['text']) == 0:
                label_1_score['bg'] = "yellow"
            else:
                label_3_score['bg'] = "yellow"
        else:
            label_1_score['bg'] = "yellow"
        return

    # check label 3
    if label_3_score['bg'] == "yellow":
        label_3_score['bg'] = "white"
        # check if label 4 is in use
        if label_4_score['text'] != "":
            if int(label_4_score['text']) == 0:
                if int(label_1_score['text']) == 0:
                    label_2_score['bg'] = "yellow"
                else:
                    label_1_score['bg'] = "yellow"
            else:
                label_4_score['bg'] = "yellow"
        else:
            if int(label_1_score['text']) == 0:
                label_2_score['bg'] = "yellow"
            else:
                label_1_score['bg'] = "yellow"

        return

    # check label 4
    if label_4_score['bg'] == "yellow":
        label_4_score['bg'] = "white"

        if int(label_1_score['text']) == 0:
            if int(label_2_score['text']) == 0:
                label_3_score['bg'] = "yellow"
            else:
                label_2_score['bg'] = "yellow"
        else:
            label_1_score['bg'] = "yellow"
        return


def next():
    """

    :return:
    """
    # check label 1
    if label_1_score['bg'] == "yellow":
        label_1_score['bg'] = "white"

        if int(label_2_score['text']) == 0:
            # is label 3 set?
            if label_3_score['text'] != "":
                if label_4_score['text'] != "":
                    if int(label_3_score['text']) == 0:
                        if int(label_4_score['text']) == 0:
                            label_1_score['bg'] = "yellow"
                        else:
                            label_4_score['bg'] = "yellow"
                    else:
                        label_3_score['bg'] = "yellow"
        else:
            label_2_score['bg'] = "yellow"
        return

    # check label 2
    if label_2_score['bg'] == "yellow":
        label_2_score['bg'] = "white"

        if label_3_score['text'] != "":
            if label_4_score['text'] != "":
                if int(label_3_score['text']) == 0:
                    if int(label_4_score['text']) == 0:
                        label_1_score['bg'] = "yellow"
                    else:
                        label_4_score['bg'] = "yellow"
                else:
                    label_3_score['bg'] = "yellow"
            elif int(label_3_score['text']) == 0:
                label_1_score['bg'] = "yellow"
            else:
                label_3_score['bg'] = "yellow"
        else:
            label_1_score['bg'] = "yellow"
        return

    # check label 3
    if label_3_score['bg'] == "yellow":
        label_3_score['bg'] = "white"
        # check if label 4 is in use
        if label_4_score['text'] != "":
            if int(label_4_score['text']) == 0:
                if int(label_1_score['text']) == 0:
                    label_2_score['bg'] = "yellow"
                else:
                    label_1_score['bg'] = "yellow"
            else:
                label_4_score['bg'] = "yellow"
        else:
            if int(label_1_score['text']) == 0:
                label_2_score['bg'] = "yellow"
            else:
                label_1_score['bg'] = "yellow"

        return

    # check label 4
    if label_4_score['bg'] == "yellow":
        label_4_score['bg'] = "white"

        if int(label_1_score['text']) == 0:
            if int(label_2_score['text']) == 0:
                label_3_score['bg'] = "yellow"
            else:
                label_2_score['bg'] = "yellow"
        else:
            label_1_score['bg'] = "yellow"
        return


def count_down():
    """
    This function counts the score down
    :return:
    """
    count_down_button.pack()
    count_down_button.pack_forget()

    button_dart_score.pack()
    button_dart_score.place(x=90, y=300, height=30, width=80)

    result = int(zwischen_label['text'])
    zwischen_label['text'] = "0"

    darts = 3

    if label_first_dart['bg'] == "yellow":
        darts = 1
    elif label_second_dart['bg'] == "yellow":
        darts = 2
    elif label_third_dart['bg'] == "yellow":
        darts = 3

    label_first_dart['bg'] = "yellow"
    label_second_dart['bg'] = "white"
    label_third_dart['bg'] = "white"

    eins = int(label_1_score['text'])
    zwei = int(label_2_score['text'])
    drei = 501
    vier = 501

    if label_3_score['text'] != "":
        drei = int(label_3_score['text'])

    if label_4_score['text'] != "":
        vier = int(label_4_score['text'])

    # label 1
    if label_1_score['bg'] == "yellow":
        current = int(label_1_score['text'])
        if result > current:
            messagebox.showinfo("Achtung", "Sie haben überworfen!")
            result2 = 0
            add_player1(result2, darts)

        elif current > result:
            current = current - result
            label_1_score['text'] = current
            add_player1(result, darts)

        elif result == current:
            current = current - result
            label_1_score['text'] = current
            add_player1(result, darts)
            eins = 0

            # first check: alle Spieler dabei
            if label_2_score['text'] != "" and label_3_score['text'] != "" and label_4_score['text'] != "":
                # check if ein spieler bereits bei 0
                if (zwei == 0 and drei != 0 and vier != 0) or (zwei != 0 and drei == 0 and vier != 0) or (
                        zwei != 0 and drei != 0 and vier == 0):
                    messagebox.showinfo("Info", label_player_1_name['text'] + " is the second winner.")

                # check if zwei Spieler bereits bei 0
                elif (zwei == 0 and drei == 0 and vier != 0) or (zwei == 0 and drei != 0 and vier == 0) or (
                        zwei != 0 and drei == 0 and vier == 0):
                    messagebox.showinfo("Info", label_player_1_name['text'] + " is the third winner.")
                    end_game()
                    return
                else:
                    messagebox.showinfo("Info", label_player_1_name['text'] + " is the first winner.")

            # second check: only 3 player
            elif label_2_score['text'] != "" and label_3_score['text'] != "" and label_4_score['text'] == "":
                # check if ein Spieler bereits bei 0
                if (zwei == 0 and drei != 0) or (zwei != 0 and drei == 0):
                    messagebox.showinfo("Info", label_player_1_name['text'] + " is the second winner.")
                    end_game()
                    return
                else:
                    messagebox.showinfo("Info", label_player_1_name['text'] + " is the first winner.")

            # third check: only 2 player
            elif label_2_score['text'] != "" and label_3_score['text'] == "" and label_4_score['text'] == "":
                messagebox.showinfo("Info", label_player_1_name['text'] + " is the first winner.")
                end_game()
                return
        else:
            messagebox.showerror("Error", "Systemerror. Bitte neustarten.")

    # label 2
    if label_2_score['bg'] == "yellow":
        current = int(label_2_score['text'])
        if result > current:
            messagebox.showinfo("Achtung", "Sie haben überworfen!")
            result2 = 0
            add_player2(result2, darts)

        elif current > result:
            current = current - result
            label_2_score['text'] = current
            add_player2(result, darts)

        elif result == current:
            current = current - result
            label_2_score['text'] = current
            add_player2(result, darts)
            zwei = 0

            # first check: alle Spieler dabei
            if label_1_score['text'] != "" and label_3_score['text'] != "" and label_4_score['text'] != "":
                # check if ein spieler bereits bei 0
                if (eins == 0 and drei != 0 and vier != 0) or (eins != 0 and drei == 0 and vier != 0) or (
                        eins != 0 and drei != 0 and vier == 0):
                    messagebox.showinfo("Info", label_player_2_name['text'] + " is the second winner.")

                # check if zwei Spieler bereits bei 0
                elif (eins == 0 and drei == 0 and vier != 0) or (eins == 0 and drei != 0 and vier == 0) or (
                        eins != 0 and drei == 0 and vier == 0):
                    messagebox.showinfo("Info", label_player_2_name['text'] + " is the third winner.")
                    end_game()
                    return
                else:
                    messagebox.showinfo("Info", label_player_2_name['text'] + " is the first winner.")

            # second check: only 3 player
            elif label_1_score['text'] != "" and label_3_score['text'] != "" and label_4_score['text'] == "":
                # check if ein Spieler bereits bei 0
                if (eins == 0 and drei != 0) or (eins != 0 and drei == 0):
                    messagebox.showinfo("Info", label_player_2_name['text'] + " is the second winner.")
                    end_game()
                    return
                else:
                    messagebox.showinfo("Info", label_player_2_name['text'] + " is the first winner.")

            # third check: only 2 player
            elif label_1_score['text'] != "" and label_3_score['text'] == "" and label_4_score['text'] == "":
                messagebox.showinfo("Info", label_player_2_name['text'] + " is the first winner.")
                end_game()
                return

        else:
            messagebox.showerror("Error", "Systemerror. Bitte neustarten.")

    # label 3
    if label_3_score['bg'] == "yellow":
        current = int(label_3_score['text'])
        if result > current:
            messagebox.showinfo("Achtung", "Sie haben überworfen!")
            result2 = 0
            add_player3(result2, darts)

        elif current > result:
            current = current - result
            label_3_score['text'] = current
            add_player3(result, darts)

        elif result == current:
            current = current - result
            label_3_score['text'] = current
            add_player3(result, darts)
            drei = 0

            # first check: alle Spieler dabei
            if label_1_score['text'] != "" and label_2_score['text'] != "" and label_4_score['text'] != "":
                # check if ein spieler bereits bei 0
                if (zwei == 0 and eins != 0 and vier != 0) or (zwei != 0 and eins == 0 and vier != 0) or (
                        zwei != 0 and eins != 0 and vier == 0):
                    messagebox.showinfo("Info", label_player_3_name['text'] + " is the second winner.")

                # check if zwei Spieler bereits bei 0
                elif (zwei == 0 and eins == 0 and vier != 0) or (zwei == 0 and eins != 0 and vier == 0) or (
                        zwei != 0 and eins == 0 and vier == 0):
                    messagebox.showinfo("Info", label_player_3_name['text'] + " is the third winner.")
                    end_game()
                    return
                else:
                    messagebox.showinfo("Info", label_player_3_name['text'] + " is the first winner.")

            # second check: only 3 player
            elif label_1_score['text'] != "" and label_2_score['text'] != "" and label_4_score['text'] == "":
                # check if ein Spieler bereits bei 0
                if (zwei == 0 and eins != 0) or (zwei != 0 and eins == 0):
                    messagebox.showinfo("Info", label_player_3_name['text'] + " is the second winner.")
                    end_game()
                    return
                else:
                    messagebox.showinfo("Info", label_player_3_name['text'] + " is the first winner.")

        else:
            messagebox.showerror("Error", "Systemerror. Bitte neustarten.")

    if label_4_score['bg'] == "yellow":
        current = int(label_4_score['text'])
        if result > current:
            messagebox.showinfo("Achtung", "Sie haben überworfen!")
            result2 = 0
            add_player4(result2, darts)

        elif current > result:
            current = current - result
            label_4_score['text'] = current
            add_player4(result, darts)

        elif result == current:
            current = current - result
            label_4_score['text'] = current
            add_player4(result, darts)

            vier = 0
            # check if ein spieler bereits bei 0
            if (zwei == 0 and drei != 0 and eins != 0) or (zwei != 0 and drei == 0 and eins != 0) or (
                    zwei != 0 and drei != 0 and eins == 0):
                messagebox.showinfo("Info", label_player_4_name['text'] + " is the second winner.")

            # check if zwei Spieler bereits bei 0
            elif (zwei == 0 and drei == 0 and vier != 0) or (zwei == 0 and drei != 0 and vier == 0) or (
                    zwei != 0 and drei == 0 and vier == 0):
                messagebox.showinfo("Info", label_player_4_name['text'] + " is the third winner.")
                end_game()
                return
            else:
                messagebox.showinfo("Info", label_player_4_name['text'] + " is the first winner.")

        else:
            messagebox.showerror("Error", "Systemerror. Bitte neustarten.")

    next()


def add():
    """

    :return:
    """
    count = int(label_dart_score['text'])
    current = int(zwischen_label['text'])
    result = current + count
    zwischen_label['text'] = result
    label_dart_score['text'] = ""

    if (label_1_score['bg'] == "yellow" and int(label_1_score['text'] == result)) \
            or (label_2_score['bg'] == "yellow" and int(label_2_score['text'] == result)) \
            or (label_3_score['bg'] == "yellow" and int(label_3_score['text'] == result)) \
            or (label_4_score['bg'] == "yellow" and int(label_4_score['text'] == result)):
        count_down_button.pack()
        count_down_button.place(x=440, y=300, height=30, width=90)
        button_dart_score.pack()
        button_dart_score.pack_forget()
        return

    if label_first_dart['bg'] == "yellow":
        label_first_dart['bg'] = "white"
        label_second_dart['bg'] = "yellow"
        return

    if label_second_dart['bg'] == "yellow":
        label_second_dart['bg'] = "white"
        label_third_dart['bg'] = "yellow"
        return

    if label_third_dart['bg'] == "yellow":
        count_down_button.pack()
        count_down_button.place(x=440, y=300, height=30, width=90)
        button_dart_score.pack()
        button_dart_score.pack_forget()


def reset():
    """

    :return:
    """
    label_1_score['text'] = "501"
    label_2_score['text'] = "501"
    label_3_score['text'] = "501"
    label_4_score['text'] = "501"
    label_player_1_name['text'] = "Player 1: "
    label_player_2_name['text'] = "Player 2: "
    label_player_3_name['text'] = "Player 3: "
    label_player_4_name['text'] = "Player 4: "

    zwischen_label['text'] = "0"
    label_dart_score['text'] = ""

    label_first_dart['bg'] = "yellow"
    label_second_dart['bg'] = "white"
    label_third_dart['bg'] = "white"

    label_1_score['bg'] = "yellow"
    label_2_score['bg'] = "white"
    label_3_score['bg'] = "white"
    label_4_score['bg'] = "white"

    save_score()

    for item in player1:
        player1.remove(item)
    for item in player2:
        player2.remove(item)
    for item in player3:
        player3.remove(item)
    for item in player4:
        player4.remove(item)


def new_game():
    """

    return:
    """
    label_1_score['bg'] = "yellow"
    label_2_score['bg'] = "white"
    label_3_score['bg'] = "white"
    label_4_score['bg'] = "white"

    if label_switch_score['text'] == "501":
        label_1_score['text'] = "501"
        label_2_score['text'] = "501"

        if label_player_3_name['text'] != "":
            label_3_score['text'] = "501"

        if label_player_4_name['text'] != "":
            label_4_score['text'] = "501"
    else:
        label_1_score['text'] = "301"
        label_2_score['text'] = "301"
        if label_player_3_name['text'] != "":
            label_3_score['text'] = "301"

        if label_player_4_name['text'] != "":
            label_4_score['text'] = "301"

    save_score()

    for item in player1:
        player1.remove(item)
    for item in player2:
        player2.remove(item)
    for item in player3:
        player3.remove(item)
    for item in player4:
        player4.remove(item)


def end_game():
    """

    """
    messagebox.showinfo("Info", "Spiel beendet.")

    label_1_score['bg'] = "yellow"
    label_2_score['bg'] = "white"
    label_3_score['bg'] = "white"
    label_4_score['bg'] = "white"

    if label_switch_score['text'] == "501":
        label_1_score['text'] = "501"
        label_2_score['text'] = "501"

        if label_player_3_name['text'] != "":
            label_3_score['text'] = "501"

        if label_player_4_name['text'] != "":
            label_4_score['text'] = "501"
    else:
        label_1_score['text'] = "301"
        label_2_score['text'] = "301"
        if label_player_3_name['text'] != "":
            label_3_score['text'] = "301"

        if label_player_4_name['text'] != "":
            label_4_score['text'] = "301"

    save_score()

    for item in player1:
        player1.remove(item)
    for item in player2:
        player2.remove(item)
    for item in player3:
        player3.remove(item)
    for item in player4:
        player4.remove(item)


if __name__ == "__main__":
    # configure the window to generate
    gui = Tk()
    gui.geometry('1275x645')
    gui.resizable(width=0, height=0)
    gui.title("Dart - User Interface")
    gui.configure(background='grey')

    # define the exit - button
    exit_button = Button(gui, text="Beenden", command=button_exit, fg="black", bg="lightgreen",
                         font=('Arial', 10, 'bold'))
    exit_button.place(x=1175, y=0, height=80, width=100)

    # ################# ---------------------- ##################
    # labels for 4 players
    label_player_1_name = Label(gui, text="Player 1: ", fg="black", font=('Arial', 13, 'bold'))
    label_player_2_name = Label(gui, text="Player 2: ", fg="black", font=('Arial', 13, 'bold'))
    label_player_3_name = Label(gui, text="Player 3: ", fg="black", font=('Arial', 13, 'bold'))
    label_player_4_name = Label(gui, text="Player 4: ", fg="black", font=('Arial', 13, 'bold'))

    label_player_1_name.place(x=10, y=10, height=30, width=110)
    label_player_2_name.place(x=160, y=10, height=30, width=110)
    label_player_3_name.place(x=310, y=10, height=30, width=110)
    label_player_4_name.place(x=460, y=10, height=30, width=110)

    # ################# ---------------------- ##################
    # text input for 4 player names and buttons
    eingabefeld_p1 = Entry(gui, bd=4, font=('Arial', 13))
    eingabefeld_p2 = Entry(gui, bd=4, font=('Arial', 13))
    eingabefeld_p3 = Entry(gui, bd=4, font=('Arial', 13))
    eingabefeld_p4 = Entry(gui, bd=4, font=('Arial', 13))

    eingabefeld_p1.place(x=750, y=5, width=100, height=30)
    eingabefeld_p2.place(x=750, y=35, width=100, height=30)
    eingabefeld_p3.place(x=900, y=5, width=100, height=30)
    eingabefeld_p4.place(x=900, y=35, width=100, height=30)

    button_name_1 = Button(gui, text="P1", bd=4, fg="black", bg="grey", font=('Arial', 10),
                           command=button_name_1)
    button_name_2 = Button(gui, text="P2", bd=4, fg="black", bg="white", font=('Arial', 10),
                           command=button_name_2)
    button_name_3 = Button(gui, text="P3", bd=4, fg="black", bg="grey", font=('Arial', 10),
                           command=button_name_3)
    button_name_4 = Button(gui, text="P4", bd=4, fg="black", bg="white", font=('Arial', 10),
                           command=button_name_4)

    button_name_1.place(x=850, y=5, height=30)
    button_name_2.place(x=850, y=35, height=30)
    button_name_3.place(x=1000, y=5, height=30)
    button_name_4.place(x=1000, y=35, height=30)

    # ################# ---------------------- ##################
    # labels for game score for 4 players
    label_1_score = Label(gui, text="501", fg="black", bg="white", font=('Arial', 13, 'bold'))
    label_2_score = Label(gui, text="501", fg="black", bg="white", font=('Arial', 13, 'bold'))
    label_3_score = Label(gui, text="501", fg="black", bg="white", font=('Arial', 13, 'bold'))
    label_4_score = Label(gui, text="501", fg="black", bg="white", font=('Arial', 13, 'bold'))

    label_1_score.place(x=10, y=60, height=30, width=110)
    label_2_score.place(x=160, y=60, height=30, width=110)
    label_3_score.place(x=310, y=60, height=30, width=110)
    label_4_score.place(x=460, y=60, height=30, width=110)

    # ################# ---------------------- ##################
    # label and button for switching points
    label_switch_score = Label(gui, text="501", fg="black", font=('Arial', 13, 'bold'))
    button_switch_score_inc = Button(gui, text="+", bd=4, fg="black", bg="lightgreen", font=('Arial', 10),
                                     command=button_switch_score_inc)
    button_switch_score_dec = Button(gui, text="-", bd=4, fg="black", bg="red", font=('Arial', 10),
                                     command=button_switch_score_dec)

    label_switch_score.place(x=1105, y=90, height=30, width=110)
    button_switch_score_dec.place(x=1215, y=90, height=30, width=30)
    button_switch_score_inc.place(x=1240, y=90, height=30, width=30)

    # ################# ---------------------- ##################
    # start - button und stop - button
    button_start_game = Button(gui, text="Start", bd=4, fg="black", bg="yellow", font=('Arial', 11),
                               command=button_start_game_function)
    button_start_game.place(x=850, y=90, height=30, width=100)

    button_stop_game = Button(gui, text="Stop", bd=4, fg="black", bg="red", font=('Arial', 11),
                              command=button_stop_game_function)
    button_stop_game.pack()
    button_stop_game.pack_forget()

    # ################# ---------------------- ##################
    # next - button
    next_button = Button(gui, text="Next", bd=4, fg="black", bg="yellow", font=('Arial', 11),
                         command=next_button)
    next_button.place(x=610, y=60, height=30, width=100)
    next_button.pack()
    next_button.pack_forget()

    # ################# ---------------------- ##################
    # eingabefeld für scoring (Punkte, die abgezogen werden sollen)
    label_dart_score = Label(gui, text="", bd=4, font=('Arial', 13))
    button_dart_score = Button(gui, text="Add", bd=4, fg="black", bg="lightgreen", font=('Arial', 10),
                               command=add)
    label_dart_score.place(x=0, y=300, height=30, width=90)
    button_dart_score.place(x=100, y=300, height=30, width=80)

    label_dart_score.pack()
    button_dart_score.pack()

    label_dart_score.pack_forget()
    button_dart_score.pack_forget()

    # ################# ---------------------- ##################
    # create Label for 1,2 and 3 Darts and zwischen and count down button
    label_first_dart = Label(gui, text="1", bd=4, bg="yellow", font=('Arial', 13))
    label_second_dart = Label(gui, text="2", bd=4, bg="white", font=('Arial', 13))
    label_third_dart = Label(gui, text="3", bd=4, bg="white", font=('Arial', 13))

    zwischen_label = Label(gui, text="0", bd=4, bg="yellow", font=('Arial', 13))
    count_down_button = Button(gui, text="Count down", bd=4, fg="black", bg="lightgreen", font=('Arial', 10),
                               command=count_down)

    label_first_dart.place(x=210, y=300, height=30, width=30)
    label_second_dart.place(x=250, y=300, height=30, width=30)
    label_third_dart.place(x=290, y=300, height=30, width=30)
    zwischen_label.place(x=340, y=300, height=30, width=100)
    count_down_button.place(x=440, y=300, height=30, width=90)

    label_first_dart.pack()
    label_second_dart.pack()
    label_third_dart.pack()
    zwischen_label.pack()
    count_down_button.pack()

    label_first_dart.pack_forget()
    label_second_dart.pack_forget()
    label_third_dart.pack_forget()
    zwischen_label.pack_forget()
    count_down_button.pack_forget()

    # ################# ---------------------- ##################
    # calculate kpis button
    button_create_excel = Button(gui, text="Score berechnen", bd=4, fg="black", bg="lightblue", font=('Arial', 11),
                                 command=create_excel)
    button_create_excel.place(x=1125, y=250, height=80, width=150)

    # ################# ---------------------- ##################
    # here set the values for all numbers (Tripel, Double and Single)
    button_triple_20 = Button(gui, text="T20", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=T20)
    button_single_20 = Button(gui, text="S20", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=S20)
    button_double_20 = Button(gui, text="D20", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=D20)

    button_triple_19 = Button(gui, text="T19", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=T19)
    button_single_19 = Button(gui, text="S19", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=S19)
    button_double_19 = Button(gui, text="D19", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=D19)

    button_triple_18 = Button(gui, text="T18", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=T18)
    button_single_18 = Button(gui, text="S18", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=S18)
    button_double_18 = Button(gui, text="D18", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=D18)

    button_triple_17 = Button(gui, text="T17", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=T17)
    button_single_17 = Button(gui, text="S17", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=S17)
    button_double_17 = Button(gui, text="D17", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=D17)

    button_triple_16 = Button(gui, text="T16", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=T16)
    button_single_16 = Button(gui, text="S16", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=S16)
    button_double_16 = Button(gui, text="D16", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=D16)
    button_triple_15 = Button(gui, text="T15", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=T15)
    button_single_15 = Button(gui, text="S15", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=S15)
    button_double_15 = Button(gui, text="D15", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=D15)
    button_triple_14 = Button(gui, text="T14", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=T14)
    button_single_14 = Button(gui, text="S14", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=S14)
    button_double_14 = Button(gui, text="D14", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=D14)

    button_triple_13 = Button(gui, text="T13", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=T13)
    button_single_13 = Button(gui, text="S13", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=S13)
    button_double_13 = Button(gui, text="D13", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=D13)
    button_triple_12 = Button(gui, text="T12", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=T12)
    button_single_12 = Button(gui, text="S12", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=S12)
    button_double_12 = Button(gui, text="D12", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=D12)
    button_triple_11 = Button(gui, text="T11", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=T11)
    button_single_11 = Button(gui, text="S11", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=S11)
    button_double_11 = Button(gui, text="D11", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=D11)

    button_triple_10 = Button(gui, text="T10", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=T10)
    button_single_10 = Button(gui, text="S10", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=S10)
    button_double_10 = Button(gui, text="D10", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=D10)
    button_triple_9 = Button(gui, text="T9", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=T9)
    button_single_9 = Button(gui, text="S9", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=S9)
    button_double_9 = Button(gui, text="D9", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=D9)
    button_triple_8 = Button(gui, text="T8", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=T8)
    button_single_8 = Button(gui, text="S8", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=S8)
    button_double_8 = Button(gui, text="D8", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=D8)

    button_triple_7 = Button(gui, text="T7", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=T7)
    button_single_7 = Button(gui, text="S7", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=S7)
    button_double_7 = Button(gui, text="D7", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=D7)
    button_triple_6 = Button(gui, text="T6", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=T6)
    button_single_6 = Button(gui, text="S6", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=S6)
    button_double_6 = Button(gui, text="D6", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=D6)
    button_triple_5 = Button(gui, text="T5", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=T5)
    button_single_5 = Button(gui, text="S5", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=S5)
    button_double_5 = Button(gui, text="D5", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=D5)

    button_triple_4 = Button(gui, text="T4", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=T4)
    button_single_4 = Button(gui, text="S4", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=S4)
    button_double_4 = Button(gui, text="D4", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=D4)
    button_triple_3 = Button(gui, text="T3", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=T3)
    button_single_3 = Button(gui, text="S3", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=S3)
    button_double_3 = Button(gui, text="D3", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=D3)
    button_triple_2 = Button(gui, text="T2", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=T2)
    button_single_2 = Button(gui, text="S2", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=S2)
    button_double_2 = Button(gui, text="D2", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=D2)

    button_triple_1 = Button(gui, text="T1", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=T1)
    button_single_1 = Button(gui, text="S1", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=S1)
    button_double_1 = Button(gui, text="D1", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=D1)
    button_single_bull = Button(gui, text="25", bd=4, fg="black", bg="green", font=('Arial', 14),
                                command=Single_Bull)
    button_bull = Button(gui, text="BULL", bd=4, fg="black", bg="red", font=('Arial', 14),
                         command=Bull)
    button_0 = Button(gui, text="0", bd=4, fg="black", bg="green", font=('Arial', 14),
                      command=null)

    button_triple_20.place(x=0, y=400, height=60, width=60)
    button_double_20.place(x=0, y=470, height=60, width=60)
    button_single_20.place(x=0, y=540, height=60, width=60)
    button_triple_19.place(x=60, y=400, height=60, width=60)
    button_double_19.place(x=60, y=470, height=60, width=60)
    button_single_19.place(x=60, y=540, height=60, width=60)
    button_triple_18.place(x=120, y=400, height=60, width=60)
    button_double_18.place(x=120, y=470, height=60, width=60)
    button_single_18.place(x=120, y=540, height=60, width=60)
    button_triple_17.place(x=180, y=400, height=60, width=60)
    button_double_17.place(x=180, y=470, height=60, width=60)
    button_single_17.place(x=180, y=540, height=60, width=60)

    button_triple_16.place(x=240, y=400, height=60, width=60)
    button_double_16.place(x=240, y=470, height=60, width=60)
    button_single_16.place(x=240, y=540, height=60, width=60)
    button_triple_15.place(x=300, y=400, height=60, width=60)
    button_double_15.place(x=300, y=470, height=60, width=60)
    button_single_15.place(x=300, y=540, height=60, width=60)
    button_triple_14.place(x=360, y=400, height=60, width=60)
    button_double_14.place(x=360, y=470, height=60, width=60)
    button_single_14.place(x=360, y=540, height=60, width=60)
    button_triple_13.place(x=420, y=400, height=60, width=60)
    button_double_13.place(x=420, y=470, height=60, width=60)
    button_single_13.place(x=420, y=540, height=60, width=60)

    button_triple_12.place(x=480, y=400, height=60, width=60)
    button_double_12.place(x=480, y=470, height=60, width=60)
    button_single_12.place(x=480, y=540, height=60, width=60)
    button_triple_11.place(x=540, y=400, height=60, width=60)
    button_double_11.place(x=540, y=470, height=60, width=60)
    button_single_11.place(x=540, y=540, height=60, width=60)
    button_triple_10.place(x=600, y=400, height=60, width=60)
    button_double_10.place(x=600, y=470, height=60, width=60)
    button_single_10.place(x=600, y=540, height=60, width=60)
    button_triple_9.place(x=660, y=400, height=60, width=60)
    button_double_9.place(x=660, y=470, height=60, width=60)
    button_single_9.place(x=660, y=540, height=60, width=60)

    button_triple_8.place(x=720, y=400, height=60, width=60)
    button_double_8.place(x=720, y=470, height=60, width=60)
    button_single_8.place(x=720, y=540, height=60, width=60)
    button_triple_7.place(x=780, y=400, height=60, width=60)
    button_double_7.place(x=780, y=470, height=60, width=60)
    button_single_7.place(x=780, y=540, height=60, width=60)
    button_triple_6.place(x=840, y=400, height=60, width=60)
    button_double_6.place(x=840, y=470, height=60, width=60)
    button_single_6.place(x=840, y=540, height=60, width=60)
    button_triple_5.place(x=900, y=400, height=60, width=60)
    button_double_5.place(x=900, y=470, height=60, width=60)
    button_single_5.place(x=900, y=540, height=60, width=60)

    button_triple_4.place(x=960, y=400, height=60, width=60)
    button_double_4.place(x=960, y=470, height=60, width=60)
    button_single_4.place(x=960, y=540, height=60, width=60)
    button_triple_3.place(x=1020, y=400, height=60, width=60)
    button_double_3.place(x=1020, y=470, height=60, width=60)
    button_single_3.place(x=1020, y=540, height=60, width=60)
    button_triple_2.place(x=1080, y=400, height=60, width=60)
    button_double_2.place(x=1080, y=470, height=60, width=60)
    button_single_2.place(x=1080, y=540, height=60, width=60)
    button_triple_1.place(x=1140, y=400, height=60, width=60)
    button_double_1.place(x=1140, y=470, height=60, width=60)
    button_single_1.place(x=1140, y=540, height=60, width=60)

    button_single_bull.place(x=1210, y=400, height=60, width=60)
    button_bull.place(x=1210, y=470, height=60, width=60)
    button_0.place(x=1210, y=540, height=60, width=60)

    # ################# ---------------------- ##################
    # reset - button
    reset_button = Button(gui, text="Reset", bd=4, fg="black", bg="red", font=('Arial', 11),
                          command=reset)

    reset_button.place(x=1175, y=190, height=30, width=100)

    new_game_button = Button(gui, text="New Game", bd=4, fg="black", bg="red", font=('Arial', 11),
                             command=new_game)

    new_game_button.place(x=850, y=200, height=30, width=100)

    new_game_button.pack()
    new_game_button.pack_forget()
    # ################# ---------------------- ##################
    gui.mainloop()
