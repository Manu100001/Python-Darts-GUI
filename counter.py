#!/usr/bin/python3
# pylint: disable=C0302
"""
This script will help you play darts.
If you don't have an electric dartboard but a normal one,
this script will help you to calculate the scores.

:author: Manuel Milde manuelmilde@gmx.net
:copyright: 2021 Manuel Milde
"""
import os
from tkinter import Tk
from tkinter import Toplevel
from tkinter import Label
from tkinter import Button
from tkinter import Entry
from tkinter import messagebox
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# color for Excel
greenFill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

# months for creating time-based excel
months = ['Januar', 'Februar', 'März', 'April', 'Mai', 'Juni', 'Juli',
          'August', 'September', 'Oktober', 'November', 'Dezember']

# safe darts and score for each player
player1 = []
player2 = []
player3 = []
player4 = []

# dictionaries for each player
player1_kpis = [{"Score": 0, "Darts": 0, "180": 0, "140": 0, "100": 0, "80": 0, "60": 0}]
player2_kpis = [{"Score": 0, "Darts": 0, "180": 0, "140": 0, "100": 0, "80": 0, "60": 0}]
player3_kpis = [{"Score": 0, "Darts": 0, "180": 0, "140": 0, "100": 0, "80": 0, "60": 0}]
player4_kpis = [{"Score": 0, "Darts": 0, "180": 0, "140": 0, "100": 0, "80": 0, "60": 0}]

player1_scores = [{"T20": 0, "T19": 0, "T18": 0, "S20": 0, "S19": 0, "S18": 0,
                   "Bull": 0, "Single_Bull": 0, "Triple": 0, "Double": 0, "No_Score": 0}]
player2_scores = [{"T20": 0, "T19": 0, "T18": 0, "S20": 0, "S19": 0, "S18": 0,
                   "Bull": 0, "Single_Bull": 0, "Triple": 0, "Double": 0, "No_Score": 0}]
player3_scores = [{"T20": 0, "T19": 0, "T18": 0, "S20": 0, "S19": 0, "S18": 0,
                   "Bull": 0, "Single_Bull": 0, "Triple": 0, "Double": 0, "No_Score": 0}]
player4_scores = [{"T20": 0, "T19": 0, "T18": 0, "S20": 0, "S19": 0, "S18": 0,
                   "Bull": 0, "Single_Bull": 0, "Triple": 0, "Double": 0, "No_Score": 0}]


def create_directory_if_not_exists():
    """
    This function creates a directory for the Excel file
    :return: return the path to the Excel file
    """
    if not os.path.isdir("Spielstände"):
        os.mkdir("Spielstände")

    if not os.path.isdir("Spielstände/Scoring"):
        os.mkdir("Spielstände/Scoring")

    # current year
    current_year = datetime.now().strftime('%Y')
    if not os.path.isdir("Spielstände/Scoring/" + current_year):
        os.mkdir("Spielstände/Scoring/" + current_year)

    # current month
    current_month = datetime.now().strftime('%m')
    month_name = months[int(current_month) - 1]
    if not os.path.isdir("Spielstände/Scoring/" + current_year + "/" + month_name):
        os.mkdir("Spielstände/Scoring/" + current_year + "/" + month_name)

    # current day
    current_day = int(datetime.now().strftime('%d'))
    date = str(current_day) + "." + str(current_month)
    if not os.path.isdir("Spielstände/Scoring/" + current_year + "/" + month_name + "/" + date):
        os.mkdir("Spielstände/Scoring/" + current_year + "/" + month_name + "/" + date)

    # create new score - file
    time = datetime.now().strftime('%H-%M-%S')

    path = "Spielstände/Scoring/" + current_year + "/" + \
           month_name + "/" + date + "/" + time + ".xlsx"

    return path


def set_standards_in_excel(sheet):
    """
    This function fills the cells in the Excel sheet
    :param sheet: the Excel sheet to write in
    :return:
    """
    sheet['B3'].fill = greenFill
    sheet['C3'].fill = greenFill
    sheet['D3'].fill = greenFill
    sheet['E3'].fill = greenFill
    sheet['F3'].fill = greenFill
    sheet['G3'].fill = greenFill
    sheet['H3'].fill = greenFill
    sheet['I3'].fill = greenFill
    sheet['J3'].fill = greenFill
    sheet['K3'].fill = redFill
    sheet['L3'].fill = redFill
    sheet['M3'].fill = greenFill
    sheet['N3'].fill = greenFill
    sheet['O3'].fill = redFill
    sheet['P3'].fill = redFill
    sheet['Q3'].fill = greenFill
    sheet['R3'].fill = redFill
    sheet['S3'].fill = greenFill
    sheet['T3'].fill = redFill
    sheet['U3'].fill = greenFill

    sheet.column_dimensions['C'].width = 8
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 18
    sheet.column_dimensions['K'].width = 14
    sheet.column_dimensions['L'].width = 14
    sheet.column_dimensions['M'].width = 14
    sheet.column_dimensions['N'].width = 14
    sheet.column_dimensions['O'].width = 14
    sheet.column_dimensions['P'].width = 14
    sheet.column_dimensions['Q'].width = 21
    sheet.column_dimensions['R'].width = 15
    sheet.column_dimensions['S'].width = 17
    sheet.column_dimensions['T'].width = 17
    sheet.column_dimensions['U'].width = 12


def fill_values_in_cells(sheet):
    """
    This function sets standard values in the Excel file
    :param sheet: the Excel sheet to write in
    :return:
    """
    sheet.cell(row=3, column=2).value = "Spieler"
    sheet.cell(row=3, column=3).value = "Average"
    sheet.cell(row=3, column=4).value = "180"
    sheet.cell(row=3, column=5).value = "140+"
    sheet.cell(row=3, column=6).value = "100+"
    sheet.cell(row=3, column=7).value = "80+"
    sheet.cell(row=3, column=8).value = "60+"

    sheet.cell(row=3, column=9).value = "Thrown Points"
    sheet.cell(row=3, column=10).value = "Thrown Darts"
    sheet.cell(row=3, column=11).value = "Thrown T20"
    sheet.cell(row=3, column=12).value = "Thrown S20"
    sheet.cell(row=3, column=13).value = "Thrown T19"
    sheet.cell(row=3, column=14).value = "Thrown S19"
    sheet.cell(row=3, column=15).value = "Thrown T18"
    sheet.cell(row=3, column=16).value = "Thrown S18"
    sheet.cell(row=3, column=17).value = "Thrown Single-Bulls"
    sheet.cell(row=3, column=18).value = "Thrown Bulls"
    sheet.cell(row=3, column=19).value = "Thrown Triple"
    sheet.cell(row=3, column=20).value = "Thrown Double"
    sheet.cell(row=3, column=21).value = "No hit"

    # set player names in Excel
    sheet.cell(row=4, column=2).value = label_player_1_name['text']
    sheet.cell(row=5, column=2).value = label_player_2_name['text']
    sheet.cell(row=6, column=2).value = label_player_3_name['text']
    sheet.cell(row=7, column=2).value = label_player_4_name['text']


def add_players(sheet):
    """
    This function adds player 3 and 4 to the Excel file if they exist
    :param sheet: the Excel sheet to write in
    :return:
    """
    # player 3
    if int(label_number_players['text']) >= 3:
        sheet.cell(row=6, column=3).value = round((player3_kpis[0]['Score'] /
                                                   player3_kpis[0]['Darts']) * 3, 2)
        sheet.cell(row=6, column=4).value = player3_kpis[0]['180']
        sheet.cell(row=6, column=5).value = player3_kpis[0]['140']
        sheet.cell(row=6, column=6).value = player3_kpis[0]['100']
        sheet.cell(row=6, column=7).value = player3_kpis[0]['80']
        sheet.cell(row=6, column=8).value = player3_kpis[0]['60']

        sheet.cell(row=6, column=9).value = player3_kpis[0]['Score']
        sheet.cell(row=6, column=10).value = player3_kpis[0]['Darts']

        sheet.cell(row=6, column=11).value = player3_scores[0]['T20']
        sheet.cell(row=6, column=12).value = player3_scores[0]['S20']
        sheet.cell(row=6, column=13).value = player3_scores[0]['T19']
        sheet.cell(row=6, column=14).value = player3_scores[0]['S19']
        sheet.cell(row=6, column=15).value = player3_scores[0]['T18']
        sheet.cell(row=6, column=16).value = player3_scores[0]['S18']
        sheet.cell(row=6, column=17).value = player3_scores[0]['Single_Bull']
        sheet.cell(row=6, column=18).value = player3_scores[0]['Bull']
        sheet.cell(row=6, column=19).value = player3_scores[0]['Triple']
        sheet.cell(row=6, column=20).value = player3_scores[0]['Double']
        sheet.cell(row=6, column=21).value = player3_scores[0]['No_Score']

    # player 4
    if int(label_number_players['text']) == 4:
        sheet.cell(row=7, column=3).value = round((player4_kpis[0]['Score'] /
                                                   player4_kpis[0]['Darts']) * 3, 2)
        sheet.cell(row=7, column=4).value = player4_kpis[0]['180']
        sheet.cell(row=7, column=5).value = player4_kpis[0]['140']
        sheet.cell(row=7, column=6).value = player4_kpis[0]['100']
        sheet.cell(row=7, column=7).value = player4_kpis[0]['80']
        sheet.cell(row=7, column=8).value = player4_kpis[0]['60']

        sheet.cell(row=7, column=9).value = player4_kpis[0]['Score']
        sheet.cell(row=7, column=10).value = player4_kpis[0]['Darts']

        sheet.cell(row=7, column=11).value = player4_scores[0]['T20']
        sheet.cell(row=7, column=12).value = player4_scores[0]['S20']
        sheet.cell(row=7, column=13).value = player4_scores[0]['T19']
        sheet.cell(row=7, column=14).value = player4_scores[0]['S19']
        sheet.cell(row=7, column=15).value = player4_scores[0]['T18']
        sheet.cell(row=7, column=16).value = player4_scores[0]['S18']
        sheet.cell(row=7, column=17).value = player4_scores[0]['Single_Bull']
        sheet.cell(row=7, column=18).value = player4_scores[0]['Bull']
        sheet.cell(row=7, column=19).value = player4_scores[0]['Triple']
        sheet.cell(row=7, column=20).value = player4_scores[0]['Double']
        sheet.cell(row=7, column=21).value = player4_scores[0]['No_Score']


def create_excel():
    """
    This function creates an Excel file
    :return:
    """
    path = create_directory_if_not_exists()

    excel_file = Workbook()
    sheet = excel_file.create_sheet('Scoring')

    set_standards_in_excel(sheet)
    fill_values_in_cells(sheet)

    # player 1
    sheet.cell(row=4, column=3).value = round((player1_kpis[0]['Score'] /
                                               player1_kpis[0]['Darts']) * 3, 2)
    sheet.cell(row=4, column=4).value = player1_kpis[0]['180']
    sheet.cell(row=4, column=5).value = player1_kpis[0]['140']
    sheet.cell(row=4, column=6).value = player1_kpis[0]['100']
    sheet.cell(row=4, column=7).value = player1_kpis[0]['80']
    sheet.cell(row=4, column=8).value = player1_kpis[0]['60']

    sheet.cell(row=4, column=9).value = player1_kpis[0]['Score']
    sheet.cell(row=4, column=10).value = player1_kpis[0]['Darts']

    sheet.cell(row=4, column=11).value = player1_scores[0]['T20']
    sheet.cell(row=4, column=12).value = player1_scores[0]['S20']
    sheet.cell(row=4, column=13).value = player1_scores[0]['T19']
    sheet.cell(row=4, column=14).value = player1_scores[0]['S19']
    sheet.cell(row=4, column=15).value = player1_scores[0]['T18']
    sheet.cell(row=4, column=16).value = player1_scores[0]['S18']
    sheet.cell(row=4, column=17).value = player1_scores[0]['Single_Bull']
    sheet.cell(row=4, column=18).value = player1_scores[0]['Bull']
    sheet.cell(row=4, column=19).value = player1_scores[0]['Triple']
    sheet.cell(row=4, column=20).value = player1_scores[0]['Double']
    sheet.cell(row=4, column=21).value = player1_scores[0]['No_Score']

    # player 2
    sheet.cell(row=5, column=3).value = round((player2_kpis[0]['Score'] /
                                               player2_kpis[0]['Darts']) * 3, 2)
    sheet.cell(row=5, column=4).value = player2_kpis[0]['180']
    sheet.cell(row=5, column=5).value = player2_kpis[0]['140']
    sheet.cell(row=5, column=6).value = player2_kpis[0]['100']
    sheet.cell(row=5, column=7).value = player2_kpis[0]['80']
    sheet.cell(row=5, column=8).value = player2_kpis[0]['60']

    sheet.cell(row=5, column=9).value = player2_kpis[0]['Score']
    sheet.cell(row=5, column=10).value = player2_kpis[0]['Darts']

    sheet.cell(row=5, column=11).value = player2_scores[0]['T20']
    sheet.cell(row=5, column=12).value = player2_scores[0]['S20']
    sheet.cell(row=5, column=13).value = player2_scores[0]['T19']
    sheet.cell(row=5, column=14).value = player2_scores[0]['S19']
    sheet.cell(row=5, column=15).value = player2_scores[0]['T18']
    sheet.cell(row=5, column=16).value = player2_scores[0]['S18']
    sheet.cell(row=5, column=17).value = player2_scores[0]['Single_Bull']
    sheet.cell(row=5, column=18).value = player2_scores[0]['Bull']
    sheet.cell(row=5, column=19).value = player2_scores[0]['Triple']
    sheet.cell(row=5, column=20).value = player2_scores[0]['Double']
    sheet.cell(row=5, column=21).value = player2_scores[0]['No_Score']

    add_players(sheet)

    # save excel - file
    excel_file.save(path)


def reset():
    """
    This function restores the original state
    :return:
    """
    clear_players()

    # disable excel button
    button_create_excel.pack()
    button_create_excel.pack_forget()

    # disable all throw buttons
    button_triple_20.pack()
    button_double_20.pack()
    button_single_20.pack()
    button_triple_19.pack()
    button_double_19.pack()
    button_single_19.pack()
    button_triple_18.pack()
    button_double_18.pack()
    button_single_18.pack()
    button_triple_17.pack()
    button_double_17.pack()
    button_single_17.pack()

    button_triple_20.pack_forget()
    button_double_20.pack_forget()
    button_single_20.pack_forget()
    button_triple_19.pack_forget()
    button_double_19.pack_forget()
    button_single_19.pack_forget()
    button_triple_18.pack_forget()
    button_double_18.pack_forget()
    button_single_18.pack_forget()
    button_triple_17.pack_forget()
    button_double_17.pack_forget()
    button_single_17.pack_forget()

    button_triple_16.pack()
    button_double_16.pack()
    button_single_16.pack()
    button_triple_15.pack()
    button_double_15.pack()
    button_single_15.pack()
    button_triple_14.pack()
    button_double_14.pack()
    button_single_14.pack()

    button_triple_16.pack_forget()
    button_double_16.pack_forget()
    button_single_16.pack_forget()
    button_triple_15.pack_forget()
    button_double_15.pack_forget()
    button_single_15.pack_forget()
    button_triple_14.pack_forget()
    button_double_14.pack_forget()
    button_single_14.pack_forget()

    reset2()


def reset2():
    """
    This function restores the original state
    :return:
    """
    button_triple_13.pack()
    button_double_13.pack()
    button_single_13.pack()
    button_triple_13.pack_forget()
    button_double_13.pack_forget()
    button_single_13.pack_forget()

    button_triple_12.pack()
    button_double_12.pack()
    button_single_12.pack()
    button_triple_11.pack()
    button_double_11.pack()
    button_single_11.pack()
    button_triple_10.pack()
    button_double_10.pack()
    button_single_10.pack()
    button_triple_9.pack()
    button_double_9.pack()
    button_single_9.pack()

    button_triple_12.pack_forget()
    button_double_12.pack_forget()
    button_single_12.pack_forget()
    button_triple_11.pack_forget()
    button_double_11.pack_forget()
    button_single_11.pack_forget()
    button_triple_10.pack_forget()
    button_double_10.pack_forget()
    button_single_10.pack_forget()
    button_triple_9.pack_forget()
    button_double_9.pack_forget()
    button_single_9.pack_forget()

    button_triple_8.pack()
    button_double_8.pack()
    button_single_8.pack()
    button_triple_7.pack()
    button_double_7.pack()
    button_single_7.pack()
    button_triple_6.pack()
    button_double_6.pack()
    button_single_6.pack()

    button_triple_8.pack_forget()
    button_double_8.pack_forget()
    button_single_8.pack_forget()
    button_triple_7.pack_forget()
    button_double_7.pack_forget()
    button_single_7.pack_forget()
    button_triple_6.pack_forget()
    button_double_6.pack_forget()
    button_single_6.pack_forget()

    reset3()


def reset3():
    """
    This function restores the original state
    :return:
    """
    button_triple_5.pack()
    button_double_5.pack()
    button_single_5.pack()
    button_triple_4.pack()
    button_double_4.pack()
    button_single_4.pack()
    button_triple_3.pack()
    button_double_3.pack()
    button_single_3.pack()
    button_triple_2.pack()
    button_double_2.pack()
    button_single_2.pack()
    button_triple_1.pack()
    button_double_1.pack()
    button_single_1.pack()

    button_triple_5.pack_forget()
    button_double_5.pack_forget()
    button_single_5.pack_forget()
    button_triple_4.pack_forget()
    button_double_4.pack_forget()
    button_single_4.pack_forget()
    button_triple_3.pack_forget()
    button_double_3.pack_forget()
    button_single_3.pack_forget()
    button_triple_2.pack_forget()
    button_double_2.pack_forget()
    button_single_2.pack_forget()
    button_triple_1.pack_forget()
    button_double_1.pack_forget()
    button_single_1.pack_forget()

    button_bull.pack()
    button_single_bull.pack()
    button_0.pack()

    button_bull.pack_forget()
    button_single_bull.pack_forget()
    button_0.pack_forget()

    # disable all player names and scores
    label_player_1_name.pack()
    label_player_2_name.pack()
    label_player_3_name.pack()
    label_player_4_name.pack()

    label_player_1_name.pack_forget()
    label_player_2_name.pack_forget()
    label_player_3_name.pack_forget()
    label_player_4_name.pack_forget()

    reset4()


def reset4():
    """
    This function restores the original state
    :return:
    """

    label_1_score.pack()
    label_2_score.pack()
    label_3_score.pack()
    label_4_score.pack()

    label_1_score.pack_forget()
    label_2_score.pack_forget()
    label_3_score.pack_forget()
    label_4_score.pack_forget()

    zwischen_label.pack()
    zwischen_label.pack_forget()

    label_first_dart.pack()
    label_second_dart.pack()
    label_third_dart.pack()

    label_first_dart.pack_forget()
    label_second_dart.pack_forget()
    label_third_dart.pack_forget()

    next_button.pack()
    next_button.pack_forget()

    button_dart_score.pack()
    button_dart_score.pack_forget()

    label_dart_score.pack()
    label_dart_score.pack_forget()

    label_switch_starting_points.place(x=587.5, y=120, height=30, width=100)
    button_dec_starting_points.place(x=557.5, y=120, height=30, width=30)
    button_inc_starting_points.place(x=687.5, y=120, height=30, width=30)

    label_number_players.place(x=587.5, y=200, height=30, width=100)
    button_minus_number_players.place(x=557.5, y=200, height=30, width=30)
    button_plus_number_players.place(x=687.5, y=200, height=30, width=30)

    input_name1.place(x=583.5, y=280, height=30, width=110)
    input_name2.place(x=583.5, y=330, height=30, width=110)

    label_welcome.place(x=337.5, y=20, height=50, width=600)
    button_continue.place(x=750, y=120, height=30, width=100)

    label_switch_starting_points['text'] = "501"
    label_number_players['text'] = "2"


def save_score():
    """
    This function saves the score for player 1 and player 2
    :return:
    """
    # update player 1
    for item in player1:
        score = item['Score']
        darts = item['Darts']

        player1_kpis[0]['Score'] += score
        player1_kpis[0]['Darts'] += darts

        if score == 180:
            player1_kpis[0]['180'] += 1

        elif score >= 140:
            player1_kpis[0]['140'] += 1

        elif score >= 100:
            player1_kpis[0]['100'] += 1

        elif score >= 80:
            player1_kpis[0]['80'] += 1

        elif score >= 60:
            player1_kpis[0]['60'] += 1

    # update player 2
    for item in player2:
        score = item['Score']
        darts = item['Darts']

        player2_kpis[0]['Score'] += score
        player2_kpis[0]['Darts'] += darts

        if score == 180:
            player2_kpis[0]['180'] += 1

        elif score >= 140:
            player2_kpis[0]['140'] += 1

        elif score >= 100:
            player2_kpis[0]['100'] += 1

        elif score >= 80:
            player2_kpis[0]['80'] += 1

        elif score >= 60:
            player2_kpis[0]['60'] += 1

    save_score2()


def save_score2():
    """
    This function is called bei save-score and saves the scores for player 3 and 4
    :return
    """
    # update player 3
    for item in player3:
        score = item['Score']
        darts = item['Darts']

        player3_kpis[0]['Score'] += score
        player3_kpis[0]['Darts'] += darts

        if score == 180:
            player3_kpis[0]['180'] += 1

        elif score >= 140:
            player3_kpis[0]['140'] += 1

        elif score >= 100:
            player3_kpis[0]['100'] += 1

        elif score >= 80:
            player3_kpis[0]['80'] += 1

        elif score >= 60:
            player3_kpis[0]['60'] += 1

    # update player 4
    for item in player4:
        score = item['Score']
        darts = item['Darts']

        player4_kpis[0]['Score'] += score
        player4_kpis[0]['Darts'] += darts

        if score == 180:
            player4_kpis[0]['180'] += 1

        elif score >= 140:
            player4_kpis[0]['140'] += 1

        elif score >= 100:
            player4_kpis[0]['100'] += 1

        elif score >= 80:
            player4_kpis[0]['80'] += 1

        elif score >= 60:
            player4_kpis[0]['60'] += 1


def add_player1(result, dart):
    """
    This function adds a throw to the list
    :param result: thrown points
    :param dart: number of darts
    :return:
    """
    player1.append({"Score": result, "Darts": dart})


def add_player2(result, dart):
    """
    This function adds a throw to the list
    :param result: thrown points
    :param dart: number of darts
    :return:
    """
    player2.append({"Score": result, "Darts": dart})


def add_player3(result, dart):
    """
    This function adds a throw to the list
    :param result: thrown points
    :param dart: number of darts
    :return:
    """
    player3.append({"Score": result, "Darts": dart})


def add_player4(result, dart):
    """
    This function adds a throw to the list
    :param result: thrown points
    :param dart: number of darts
    :return:
    """
    player4.append({"Score": result, "Darts": dart})


def t20():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T20"
    label_dart_score['text'] = "60"


def d20():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D20"
    label_dart_score['text'] = "40"


def s20():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S20"
    label_dart_score['text'] = "20"


def t19():
    """"
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T19"
    label_dart_score['text'] = "57"


def d19():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D19"
    label_dart_score['text'] = "38"


def s19():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S19"
    label_dart_score['text'] = "19"


def t18():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T18"
    label_dart_score['text'] = "54"


def d18():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D18"

    label_dart_score['text'] = "36"


def s18():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S18"

    label_dart_score['text'] = "18"


def t17():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T17"
    label_dart_score['text'] = "51"


def d17():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D17"
    label_dart_score['text'] = "34"


def s17():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S17"
    label_dart_score['text'] = "17"


def t16():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T16"
    label_dart_score['text'] = "48"


def d16():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D16"
    label_dart_score['text'] = "32"


def s16():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S16"
    label_dart_score['text'] = "16"


def t15():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T15"
    label_dart_score['text'] = "45"


def d15():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D15"
    label_dart_score['text'] = "30"


def s15():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S15"
    label_dart_score['text'] = "15"


def t14():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T14"
    label_dart_score['text'] = "42"


def d14():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D14"
    label_dart_score['text'] = "28"


def s14():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S14"
    label_dart_score['text'] = "14"


def t13():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T13"
    label_dart_score['text'] = "39"


def d13():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D13"
    label_dart_score['text'] = "26"


def s13():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S13"
    label_dart_score['text'] = "13"


def t12():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T12"
    label_dart_score['text'] = "36"


def d12():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D12"
    label_dart_score['text'] = "24"


def s12():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S12"
    label_dart_score['text'] = "12"


def t11():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T11"
    label_dart_score['text'] = "33"


def d11():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D11"
    label_dart_score['text'] = "22"


def s11():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S11"
    label_dart_score['text'] = "11"


def t10():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T10"
    label_dart_score['text'] = "30"


def d10():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D10"
    label_dart_score['text'] = "20"


def s10():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S10"
    label_dart_score['text'] = "10"


def t_9():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T9"
    label_dart_score['text'] = "27"


def d_9():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D9"
    label_dart_score['text'] = "18"


def s_9():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S9"
    label_dart_score['text'] = "9"


def t_8():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T8"
    label_dart_score['text'] = "24"


def d_8():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D8"
    label_dart_score['text'] = "16"


def s_8():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S8"
    label_dart_score['text'] = "8"


def t_7():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T7"
    label_dart_score['text'] = "21"


def d_7():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D7"
    label_dart_score['text'] = "14"


def s_7():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S7"
    label_dart_score['text'] = "7"


def t_6():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T6"
    label_dart_score['text'] = "18"


def d_6():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D6"
    label_dart_score['text'] = "12"


def s_6():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S6"
    label_dart_score['text'] = "6"


def t_5():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T5"
    label_dart_score['text'] = "15"


def d_5():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D5"
    label_dart_score['text'] = "10"


def s_5():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S5"
    label_dart_score['text'] = "5"


def t_4():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T4"
    label_dart_score['text'] = "12"


def d_4():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D4"
    label_dart_score['text'] = "8"


def s_4():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S4"
    label_dart_score['text'] = "4"


def t_3():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T3"
    label_dart_score['text'] = "9"


def d_3():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D3"
    label_dart_score['text'] = "6"


def s_3():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S3"
    label_dart_score['text'] = "3"


def t_2():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T2"
    label_dart_score['text'] = "6"


def d_2():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D2"
    label_dart_score['text'] = "4"


def s_2():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S2"
    label_dart_score['text'] = "2"


def t_1():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "T1"
    label_dart_score['text'] = "3"


def d_1():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "D1"
    label_dart_score['text'] = "2"


def s_1():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "S1"
    label_dart_score['text'] = "1"


def bull():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "Bull"
    label_dart_score['text'] = "50"


def single_bull():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = "Single_Bull"
    label_dart_score['text'] = "25"


def null():
    """
    This function adds the thrown score
    :return:
    """
    label_invisible['text'] = ""
    label_dart_score['text'] = "0"


def button_exit():
    """
    This function creates an exit - button for the gui
    :return:
    """
    if not any(isinstance(window, Toplevel) for window in gui.winfo_children()):
        exit_window = Toplevel(gui)
        exit_window.geometry('250x150')
        exit_window.resizable(width=False, height=False)
        exit_window.title("Stop?")

        label_exit = Label(exit_window, text="End game?", font=('Arial', 11))
        button_yes = Button(exit_window, text="Yes", command=exit_window.quit,
                            font=('Arial', 10, 'bold'), bg="white",
                            fg="green")
        button_no = Button(exit_window, text="No", command=exit_window.destroy,
                           font=('Arial', 10, 'bold'),
                           bg="white", fg="red")

        label_exit.place(x=80, y=0, width=100, height=50)
        button_yes.place(x=50, y=60, width=50, height=50)
        button_no.place(x=150, y=60, width=50, height=50)

    else:
        messagebox.showinfo("Info", "You already clicked on \"Stop\"!")


def increment_starting_points():
    """
    Ths function sets all labels to 501
    :return:
    """
    label_switch_starting_points['text'] = "501"


def decrement_starting_points():
    """
    This function sets all labels to 301
    :return:
    """
    label_switch_starting_points['text'] = "301"


def next_button():
    """
    This function switches to the next player
    :return:
    """
    zwischen_label['text'] = "0"
    label_first_dart['bg'] = "yellow"
    label_second_dart['bg'] = "white"
    label_third_dart['bg'] = "white"

    number = int(label_number_players['text'])

    # check label 1
    if label_1_score['bg'] == "yellow":
        label_1_score['bg'] = "white"

        if number == 2:
            label_2_score['bg'] = "yellow"

        elif number == 3:
            if int(label_2_score['text']) == 0:
                label_3_score['bg'] = "yellow"
            else:
                label_2_score['bg'] = "yellow"

        elif number == 4:
            if int(label_2_score['text']) == 0:
                if int(label_3_score['text']) == 0:
                    label_4_score['bg'] = "yellow"
                else:
                    label_3_score['bg'] = "yellow"
            else:
                label_2_score['bg'] = "yellow"
    else:
        next_button2(number)


def next_button2(number):
    """
    This function switches to the next player
    :param number: number of players
    :return:
    """
    # check label 2
    if label_2_score['bg'] == "yellow":
        label_2_score['bg'] = "white"

        if number == 2:
            label_1_score['bg'] = "yellow"

        elif number == 3:
            if int(label_3_score['text']) == 0:
                label_1_score['bg'] = "yellow"
            else:
                label_3_score['bg'] = "yellow"

        elif number == 4:
            if int(label_3_score['text']) == 0:
                if int(label_4_score['text']) == 0:
                    label_1_score['bg'] = "yellow"
                else:
                    label_4_score['bg'] = "yellow"
            else:
                label_3_score['bg'] = "yellow"

    else:
        next_button3(number)


def next_button3(number):
    """
    This function switches to the next player
    :param number: number of players
    :return:
    """
    # check label 3
    if label_3_score['bg'] == "yellow":
        label_3_score['bg'] = "white"

        if number == 3:
            if int(label_1_score['text']) == 0:
                label_2_score['bg'] = "yellow"
            else:
                label_1_score['bg'] = "yellow"

        elif number == 4:
            if int(label_4_score['text']) == 0:
                if int(label_1_score['text']) == 0:
                    label_2_score['bg'] = "yellow"
                else:
                    label_1_score['bg'] = "yellow"
            else:
                label_4_score['bg'] = "yellow"
    else:
        next_button4(number)


def next_button4(number):
    """
    This function switches to the next player
    :param number: number of players
    :return:
    """
    # check label 4
    if label_4_score['bg'] == "yellow":
        label_4_score['bg'] = "white"

        if number == 4:
            if int(label_1_score['text']) == 0:
                if int(label_2_score['text']) == 0:
                    label_3_score['bg'] = "yellow"
                else:
                    label_2_score['bg'] = "yellow"
            else:
                label_1_score['bg'] = "yellow"


def next_label():
    """
    This function switches to the next player
    :return:
    """
    number = int(label_number_players['text'])

    # check label 1
    if label_1_score['bg'] == "yellow":
        label_1_score['bg'] = "white"

        if number == 2:
            label_2_score['bg'] = "yellow"

        elif number == 3:
            if int(label_2_score['text']) == 0:
                label_3_score['bg'] = "yellow"
            else:
                label_2_score['bg'] = "yellow"

        elif number == 4:
            if int(label_2_score['text']) == 0:
                if int(label_3_score['text']) == 0:
                    label_4_score['bg'] = "yellow"
                else:
                    label_3_score['bg'] = "yellow"
            else:
                label_2_score['bg'] = "yellow"
    else:
        next_label2(number)


def next_label2(number):
    """
    This function switches to the next player
    :param number: number of players
    :return:
    """
    # check label 2
    if label_2_score['bg'] == "yellow":
        label_2_score['bg'] = "white"

        if number == 2:
            label_1_score['bg'] = "yellow"

        elif number == 3:
            if int(label_3_score['text']) == 0:
                label_1_score['bg'] = "yellow"
            else:
                label_3_score['bg'] = "yellow"

        elif number == 4:
            if int(label_3_score['text']) == 0:
                if int(label_4_score['text']) == 0:
                    label_1_score['bg'] = "yellow"
                else:
                    label_4_score['bg'] = "yellow"
            else:
                label_3_score['bg'] = "yellow"

    else:
        next_label3(number)


def next_label3(number):
    """
    This function switches to the next player
    :param number: number of players
    :return:
    """
    # check label 3
    if label_3_score['bg'] == "yellow":
        label_3_score['bg'] = "white"

        if number == 3:
            if int(label_1_score['text']) == 0:
                label_2_score['bg'] = "yellow"
            else:
                label_1_score['bg'] = "yellow"

        elif number == 4:
            if int(label_4_score['text']) == 0:
                if int(label_1_score['text']) == 0:
                    label_2_score['bg'] = "yellow"
                else:
                    label_1_score['bg'] = "yellow"
            else:
                label_4_score['bg'] = "yellow"
    else:
        next_label4(number)


def next_label4(number):
    """
    This function switches to the next player
    :param number: number of players
    :return:
    """
    # check label 4
    if label_4_score['bg'] == "yellow":
        label_4_score['bg'] = "white"

        if number == 4:
            if int(label_1_score['text']) == 0:
                if int(label_2_score['text']) == 0:
                    label_3_score['bg'] = "yellow"
                else:
                    label_2_score['bg'] = "yellow"
            else:
                label_1_score['bg'] = "yellow"


def get_amount_of_darts():
    """
    This function calculate the number of thrown darts
    :return: amount of darts
    """
    count_down_button.pack()
    count_down_button.pack_forget()
    count_down_button['text'] = "Count down"

    button_dart_score.pack()
    button_dart_score.place(x=90, y=300, height=30, width=80)

    zwischen_label['text'] = "0"

    darts = 3

    if label_first_dart['bg'] == "yellow":
        darts = 1
    elif label_second_dart['bg'] == "yellow":
        darts = 2
    elif label_third_dart['bg'] == "yellow":
        darts = 3

    label_first_dart['bg'] = "yellow"
    label_second_dart['bg'] = "white"
    label_third_dart['bg'] = "white"

    return darts


def count_down():
    """
    This function counts the score down
    :return:
    """
    number = int(label_number_players['text'])
    result = int(zwischen_label['text'])
    darts = get_amount_of_darts()

    two = int(label_2_score['text'])

    try:
        three = int(label_3_score['text'])
        four = int(label_4_score['text'])

    except ValueError:
        three = 501
        four = 501

    # count down player 1
    if label_1_score['bg'] == "yellow":
        current = int(label_1_score['text'])
        if result > current:
            messagebox.showinfo("Warning", "No score.")
            result2 = 0
            add_player1(result2, darts)

        elif current > result:
            current = current - result
            label_1_score['text'] = current
            add_player1(result, darts)

        elif result == current:
            current = current - result
            label_1_score['text'] = current
            add_player1(result, darts)

            # first check: 2 players
            if number == 2:
                messagebox.showinfo("Info", label_player_1_name['text'] + " is the winner.")
                end_game()

            # second check: 3 players
            elif number == 3:
                # check if one player has 0 points left
                if two == 0 or three == 0:
                    messagebox.showinfo("Info", label_player_1_name['text'] +
                                        " is the second winner.")
                    end_game()
                else:
                    messagebox.showinfo("Info", label_player_1_name['text'] +
                                        " is the first winner.")

            # third check: 4 players
            elif number == 4:
                check_4players_label1(two, three, four)

        else:
            messagebox.showerror("Error", "Systemerror. Please restart.")

        next_label()
    else:
        count_down_player2(result, darts)


def check_4players_label1(two, three, four):
    """
    This function check 4 players for label 1
    :param two: score of player 2
    :param three: score of player 3
    :param four: score of player 4
    :return:
    """
    players_with_zero_points = 0

    if two == 0:
        players_with_zero_points += 1

    if three == 0:
        players_with_zero_points += 1

    if four == 0:
        players_with_zero_points += 1

    if players_with_zero_points == 0:
        messagebox.showinfo("Info", label_player_1_name['text'] +
                            " is the first winner.")

    elif players_with_zero_points == 1:
        messagebox.showinfo("Info", label_player_1_name['text'] +
                            " is second the winner.")

    elif players_with_zero_points == 2:
        messagebox.showinfo("Info", label_player_1_name['text'] +
                            " is the third winner.")
        end_game()


def count_down_player2(result, darts):
    """
    This function counts the score of player 2 down
    :param result: thrown points
    :param darts: number of darts
    :return:
    """
    one = int(label_1_score['text'])

    try:
        three = int(label_3_score['text'])
        four = int(label_4_score['text'])

    except ValueError:
        three = 501
        four = 501

    number = int(label_number_players['text'])
    # count down label 2
    if label_2_score['bg'] == "yellow":
        current = int(label_2_score['text'])
        if result > current:
            messagebox.showinfo("Warning", "No score.")
            result2 = 0
            add_player2(result2, darts)

        elif current > result:
            current = current - result
            label_2_score['text'] = current
            add_player2(result, darts)

        elif result == current:
            current = current - result
            label_2_score['text'] = current
            add_player2(result, darts)

            # first check: 2 players
            if number == 2:
                messagebox.showinfo("Info", label_player_2_name['text'] + " is the winner.")
                end_game()

            # second check: 3 players
            elif number == 3:
                # check if one player has 0 points left
                if one == 0 or three == 0:
                    messagebox.showinfo("Info", label_player_2_name['text'] +
                                        " is the second winner.")
                    end_game()
                else:
                    messagebox.showinfo("Info", label_player_2_name['text'] +
                                        " is the first winner.")

            # third check: 4 players
            elif number == 4:
                check_4players_label2(one, three, four)

        else:
            messagebox.showerror("Error", "Systemerror. Bitte neustarten.")

        next_label()
    else:
        count_down_player3(result, darts)


def check_4players_label2(one, three, four):
    """
    This function check 4 players for label 2
     :param one: score of player 1
    :param three: score of player 3
    :param four: score of player 4
    :return:
    """
    players_with_zero_points = 0

    if one == 0:
        players_with_zero_points += 1

    if three == 0:
        players_with_zero_points += 1

    if four == 0:
        players_with_zero_points += 1

    if players_with_zero_points == 0:
        messagebox.showinfo("Info", label_player_2_name['text'] +
                            " is the first winner.")

    elif players_with_zero_points == 1:
        messagebox.showinfo("Info", label_player_2_name['text'] +
                            " is second the winner.")

    elif players_with_zero_points == 2:
        messagebox.showinfo("Info", label_player_2_name['text'] +
                            " is the third winner.")
        end_game()
        return


def count_down_player3(result, darts):
    """
    This function counts the score of player 3 down
    :param result: thrown points
    :param darts: number of darts
    :return:
    """
    one = int(label_1_score['text'])
    two = int(label_2_score['text'])

    try:
        four = int(label_4_score['text'])

    except ValueError:
        four = 501

    number = int(label_number_players['text'])

    # label 3
    if label_3_score['bg'] == "yellow":
        current = int(label_3_score['text'])
        if result > current:
            messagebox.showinfo("Warning", "No score.")
            result2 = 0
            add_player3(result2, darts)

        elif current > result:
            current = current - result
            label_3_score['text'] = current
            add_player3(result, darts)

        elif result == current:
            current = current - result
            label_3_score['text'] = current
            add_player3(result, darts)

            # first check: 3 players
            if number == 3:
                # check if one player has 0 points left
                if one == 0 or two == 0:
                    messagebox.showinfo("Info", label_player_3_name['text'] +
                                        " is the second winner.")
                    end_game()
                else:
                    messagebox.showinfo("Info", label_player_3_name['text'] +
                                        " is the first winner.")

            # second check: 4 players
            elif number == 4:
                check_4players_label3(one, two, four)

        else:
            messagebox.showerror("Error", "Systemerror. Bitte neustarten.")

        next_label()
    else:
        count_down_player4(result, darts)


def check_4players_label3(one, two, four):
    """
    This function check 4 players for label 3
     :param one: score of player 1
    :param two: score of player 2
    :param four: score of player 4
    :return:
    """
    players_with_zero_points = 0

    if one == 0:
        players_with_zero_points += 1

    if two == 0:
        players_with_zero_points += 1

    if four == 0:
        players_with_zero_points += 1

    if players_with_zero_points == 0:
        messagebox.showinfo("Info", label_player_3_name['text'] +
                            " is the first winner.")

    elif players_with_zero_points == 1:
        messagebox.showinfo("Info", label_player_3_name['text'] +
                            " is second the winner.")

    elif players_with_zero_points == 2:
        messagebox.showinfo("Info", label_player_3_name['text'] +
                            " is the third winner.")
        end_game()
        return


def count_down_player4(result, darts):
    """
    This function counts the score of player 3 down
    :param result: thrown points
    :param darts: number of darts
    :return:
    """
    one = int(label_1_score['text'])
    two = int(label_2_score['text'])

    try:
        three = int(label_3_score['text'])

    except ValueError:
        three = 501

    number = int(label_number_players['text'])

    if label_4_score['bg'] == "yellow":
        current = int(label_4_score['text'])
        if result > current:
            messagebox.showinfo("Warning", "No score.")
            result2 = 0
            add_player4(result2, darts)

        elif current > result:
            current = current - result
            label_4_score['text'] = current
            add_player4(result, darts)

        elif result == current:
            current = current - result
            label_4_score['text'] = current
            add_player4(result, darts)

            # first check: 4 players
            if number == 4:
                check_4players_label4(one, two, three)

        else:
            messagebox.showerror("Error", "Systemerror. Bitte neustarten.")
        next_label()


def check_4players_label4(one, two, three):
    """
    This function check 4 players for label 4
    :param one: score of player 1
    :param two: score of player 2
    :param three: score of player 3
    :return:
    """
    players_with_zero_points = 0

    if one == 0:
        players_with_zero_points += 1

    if two == 0:
        players_with_zero_points += 1

    if three == 0:
        players_with_zero_points += 1

    if players_with_zero_points == 0:
        messagebox.showinfo("Info", label_player_4_name['text'] +
                            "is the first winner.")

    elif players_with_zero_points == 1:
        messagebox.showinfo("Info", label_player_4_name['text'] +
                            " is second the winner.")

    elif players_with_zero_points == 2:
        messagebox.showinfo("Info", label_player_4_name['text'] +
                            " is the third winner.")
        end_game()
        return


def add_scores():
    """
    This function adds the thrown darts to the kpis/player_scores of player 1
    :return:
    """
    # get the score from invisible label and reset it to 0
    score = label_invisible['text']
    label_invisible['text'] = "0"

    # player 1
    if label_1_score['bg'] == "yellow":
        if "T" in score:
            player1_scores[0]["Triple"] += 1
            if score == "T20":
                player1_scores[0]["T20"] += 1
            elif score == "T19":
                player1_scores[0]["T19"] += 1
            elif score == "T18":
                player1_scores[0]["T18"] += 1

        elif "D" in score:
            player1_scores[0]["Double"] += 1

        elif "Bull" in score:
            if score == "Single_Bull":
                player1_scores[0]["Single_Bull"] += 1
            else:
                player1_scores[0]["Bull"] += 1
        elif "S" in score:
            add_scores_player1_help(score)
        else:
            player1_scores[0]["No_Score"] += 1

    else:
        add_scores_player2(score)


def add_scores_player1_help(score):
    """
    This function adds the thrown darts to the kpis/player_scores of player 1
    :param score: current score
    :return:
    """
    if score == "S20":
        player1_scores[0]["S20"] += 1
    elif score == "S19":
        player1_scores[0]["S19"] += 1
    elif score == "S18":
        player1_scores[0]["S18"] += 1


def add_scores_player2(score):
    """
    This function adds the thrown darts to the kpis/player_scores of player 2
    :param score: current score
    :return
    """
    # player 2
    if label_2_score['bg'] == "yellow":
        if "T" in score:
            player2_scores[0]["Triple"] += 1
            if score == "T20":
                player2_scores[0]["T20"] += 1
            elif score == "T19":
                player2_scores[0]["T19"] += 1
            elif score == "T18":
                player2_scores[0]["T18"] += 1

        elif "D" in score:
            player2_scores[0]["Double"] += 1

        elif "Bull" in score:
            if score == "Single_Bull":
                player2_scores[0]["Single_Bull"] += 1
            else:
                player2_scores[0]["Bull"] += 1

        elif "S" in score:
            add_scores_player2_help(score)

        else:
            player2_scores[0]["No_Score"] += 1

    else:
        add_scores_player3(score)


def add_scores_player2_help(score):
    """
    This function adds the thrown darts to the kpis/player_scores of player 2
    :param score: current score
    :return:
    """
    if score == "S20":
        player2_scores[0]["S20"] += 1
    elif score == "S19":
        player2_scores[0]["S19"] += 1
    elif score == "S18":
        player2_scores[0]["S18"] += 1


def add_scores_player3(score):
    """
    This function adds the thrown darts to the kpis/player_scores of player 3
    :param score: current score
    :return
    """
    # player 3
    if label_3_score['bg'] == "yellow":
        if "T" in score:
            player3_scores[0]["Triple"] += 1
            if score == "T20":
                player3_scores[0]["T20"] += 1
            elif score == "T19":
                player3_scores[0]["T19"] += 1
            elif score == "T18":
                player3_scores[0]["T18"] += 1

        elif "D" in score:
            player3_scores[0]["Double"] += 1

        elif "Bull" in score:
            if score == "Single_Bull":
                player3_scores[0]["Single_Bull"] += 1
            else:
                player3_scores[0]["Bull"] += 1

        elif "S" in score:
            add_scores_player3_help(score)

        else:
            player3_scores[0]["No_Score"] += 1

    else:
        add_scores_player4(score)


def add_scores_player3_help(score):
    """
    This function adds the thrown darts to the kpis/player_scores of player 3
    :param score: current score
    :return:
    """
    if score == "S20":
        player3_scores[0]["S20"] += 1
    elif score == "S19":
        player3_scores[0]["S19"] += 1
    elif score == "S18":
        player3_scores[0]["S18"] += 1


def add_scores_player4(score):
    """
    This function adds the thrown darts to the kpis/player_scores of player 4
    :param score: current score
    :return:
    """
    # player 4
    if label_4_score['bg'] == "yellow":
        if "T" in score:
            player4_scores[0]["Triple"] += 1
            if score == "T20":
                player4_scores[0]["T20"] += 1
            elif score == "T19":
                player4_scores[0]["T19"] += 1
            elif score == "T18":
                player4_scores[0]["T18"] += 1

        elif "D" in score:
            player4_scores[0]["Double"] += 1

        elif "Bull" in score:
            if score == "Single_Bull":
                player4_scores[0]["Single_Bull"] += 1
            else:
                player4_scores[0]["Bull"] += 1

        elif "S" in score:
            add_scores_player4_help(score)

        else:
            player4_scores[0]["No_Score"] += 1

    else:
        messagebox.showinfo("Error", "Restart")


def add_scores_player4_help(score):
    """
    This function adds the thrown darts to the kpis/player_scores of player 4
    :param score: current score
    :return:
    """
    if score == "S20":
        player4_scores[0]["S20"] += 1
    elif score == "S19":
        player4_scores[0]["S19"] += 1
    elif score == "S18":
        player4_scores[0]["S18"] += 1


def add():
    """
    This function adds the thrown score to the total score
    :return:
    """
    # save scores to kpis
    add_scores()

    count = int(label_dart_score['text'])
    current = int(zwischen_label['text'])

    result = current + count
    zwischen_label['text'] = result
    label_dart_score['text'] = ""

    flag1 = False

    if (label_1_score['bg'] == "yellow" and result == int(label_1_score['text'])) \
            or (label_2_score['bg'] == "yellow" and result == int(label_2_score['text'])):
        flag1 = True

    if flag1 or (label_3_score['bg'] == "yellow" and result == int(label_3_score['text'])) \
                or (label_4_score['bg'] == "yellow" and result == int(label_4_score['text'])):
        count_down_button.pack()
        count_down_button['text'] = "Count down"
        count_down_button.place(x=440, y=300, height=30, width=90)
        button_dart_score.pack()
        button_dart_score.pack_forget()
        return

    flag2 = False

    if (label_1_score['bg'] == "yellow" and result > int(label_1_score['text'])) \
            or (label_2_score['bg'] == "yellow" and result > int(label_2_score['text'])):
        flag2 = True

    if flag2 or (label_3_score['bg'] == "yellow" and result > int(label_3_score['text'])) \
                or (label_4_score['bg'] == "yellow" and result > int(label_4_score['text'])):
        count_down_button.pack()
        count_down_button['text'] = "No score. Next Player"
        count_down_button.place(x=440, y=300, height=30, width=150)
        button_dart_score.pack()
        button_dart_score.pack_forget()
        return

    if label_first_dart['bg'] == "yellow":
        label_first_dart['bg'] = "white"
        label_second_dart['bg'] = "yellow"
        return

    if label_second_dart['bg'] == "yellow":
        label_second_dart['bg'] = "white"
        label_third_dart['bg'] = "yellow"
        return

    if label_third_dart['bg'] == "yellow":
        count_down_button.pack()
        count_down_button.place(x=440, y=300, height=30, width=90)
        button_dart_score.pack()
        button_dart_score.pack_forget()


def end_game():
    """
    This function stops the game
    """
    messagebox.showinfo("Info", "Game is over.")

    label_1_score['bg'] = "yellow"
    label_2_score['bg'] = "white"
    label_3_score['bg'] = "white"
    label_4_score['bg'] = "white"

    if label_switch_starting_points['text'] == "501":
        label_1_score['text'] = "501"
        label_2_score['text'] = "501"

        if int(label_number_players['text']) == 3:
            label_3_score['text'] = "501"

        if int(label_number_players['text']) == 4:
            label_4_score['text'] = "501"
    else:
        label_1_score['text'] = "301"
        label_2_score['text'] = "301"
        if int(label_number_players['text']) == 3:
            label_3_score['text'] = "301"

        if int(label_number_players['text']) == 4:
            label_4_score['text'] = "301"

    button_create_excel.place(x=1125, y=210, height=80, width=150)
    save_score()
    clear_players()


def clear_players():
    """
    This function removes all items from the players
    :return:
    """

    while len(player1) > 0:
        for item in player1:
            player1.remove(item)

    while len(player2) > 0:
        for item in player2:
            player2.remove(item)

    while len(player3) > 0:
        for item in player3:
            player3.remove(item)

    while len(player4) > 0:
        for item in player4:
            player4.remove(item)


def plus_player_numbers():
    """
    This function increments the number of players, max. 4 players
    :return:
    """
    number = int(label_number_players['text'])

    if number == 2:
        label_number_players['text'] = 3
        input_name3.place(x=583.5, y=380, height=30, width=110)

    elif number == 3:
        label_number_players['text'] = 4
        input_name4.place(x=583.5, y=430, height=30, width=110)


def minus_player_numbers():
    """
    This function decrements the number of players, min. 2 players
    :return:
    """
    number = int(label_number_players['text'])

    if number == 4:
        label_number_players['text'] = 3
        input_name4.pack()
        input_name4.pack_forget()

    elif number == 3:
        label_number_players['text'] = 2
        input_name3.pack()
        input_name3.pack_forget()


def check_names():
    """
    This function checks if all names were entered, and then the game will start
    :return:
    """
    flag = True
    number = int(label_number_players['text'])
    name1 = input_name1.get()
    name2 = input_name2.get()
    name3 = input_name3.get()
    name4 = input_name4.get()

    if number == 2:
        if name1 == "" or name2 == "":
            messagebox.showinfo("Error", "Not all names were entered.")
            flag = False

    elif number == 3:
        if name1 == "" or name2 == "" or name3 == "":
            messagebox.showinfo("Error", "Not all names were entered.")
            flag = False
        else:
            label_player_3_name['text'] = name3

    elif number == 4:
        if name1 == "" or name2 == "" or name3 == "" or name4 == "":
            messagebox.showinfo("Error", "Not all names were entered.")
            flag = False
        else:
            label_player_3_name['text'] = name3
            label_player_4_name['text'] = name4

    if flag:
        check_names2(name1, name2)


def check_names2(name1, name2):
    """
     This function checks if all names were entered, and then the game will start
    :return:
    """
    label_player_1_name['text'] = name1
    label_player_2_name['text'] = name2

    # disable input fields and labels and buttons for selecting player numbers
    input_name1.pack()
    input_name2.pack()
    input_name3.pack()
    input_name4.pack()

    input_name1.pack_forget()
    input_name2.pack_forget()
    input_name3.pack_forget()
    input_name4.pack_forget()

    button_continue.pack()
    button_continue.pack_forget()

    # disable plus and minus
    button_plus_number_players.pack()
    button_minus_number_players.pack()

    button_plus_number_players.pack_forget()
    button_minus_number_players.pack_forget()

    label_number_players.pack()
    label_welcome.pack()

    label_number_players.pack_forget()
    label_welcome.pack_forget()

    button_inc_starting_points.pack()
    button_dec_starting_points.pack()
    label_switch_starting_points.pack()

    button_inc_starting_points.pack_forget()
    button_dec_starting_points.pack_forget()
    label_switch_starting_points.pack_forget()

    start_game()


def start_game():
    """
    This function starts the game after the names were entered
    """
    number_players = int(label_number_players['text'])
    clear_players()

    reset_button.place(x=1175, y=80, height=30, width=100)

    label_1_score['bg'] = "yellow"
    label_2_score['bg'] = "white"
    label_3_score['bg'] = "white"
    label_4_score['bg'] = "white"

    if label_switch_starting_points['text'] == "501":
        label_1_score['text'] = "501"
        label_2_score['text'] = "501"
        label_3_score['text'] = "501"
        label_4_score['text'] = "501"

    else:
        label_1_score['text'] = "301"
        label_2_score['text'] = "301"
        label_3_score['text'] = "301"
        label_4_score['text'] = "301"

    label_1_score.place(x=10, y=60, height=30, width=150)
    label_2_score.place(x=200, y=60, height=30, width=150)

    label_player_1_name.place(x=10, y=10, height=30, width=150)
    label_player_2_name.place(x=200, y=10, height=30, width=150)

    if number_players >= 3:
        label_3_score.place(x=390, y=60, height=30, width=150)
        label_player_3_name.place(x=390, y=10, height=30, width=150)

    if number_players == 4:
        label_4_score.place(x=580, y=60, height=30, width=150)
        label_player_4_name.place(x=580, y=10, height=30, width=150)

    next_button.place(x=810, y=60, height=30, width=130)

    label_dart_score.pack()
    button_dart_score.pack()

    label_dart_score.place(x=0, y=300, height=30, width=90)
    button_dart_score.place(x=90, y=300, height=30, width=80)

    label_first_dart.pack()
    label_second_dart.pack()
    label_third_dart.pack()
    zwischen_label.pack()

    label_first_dart.place(x=210, y=300, height=30, width=30)
    label_second_dart.place(x=250, y=300, height=30, width=30)
    label_third_dart.place(x=290, y=300, height=30, width=30)
    zwischen_label.place(x=340, y=300, height=30, width=100)

    enable_throw_buttons()


def enable_throw_buttons():
    """
    This function enables button (0 to T20)
    :return:
    """
    button_triple_20.place(x=0, y=400, height=60, width=60)
    button_double_20.place(x=0, y=470, height=60, width=60)
    button_single_20.place(x=0, y=540, height=60, width=60)
    button_triple_19.place(x=60, y=400, height=60, width=60)
    button_double_19.place(x=60, y=470, height=60, width=60)
    button_single_19.place(x=60, y=540, height=60, width=60)
    button_triple_18.place(x=120, y=400, height=60, width=60)
    button_double_18.place(x=120, y=470, height=60, width=60)
    button_single_18.place(x=120, y=540, height=60, width=60)
    button_triple_17.place(x=180, y=400, height=60, width=60)
    button_double_17.place(x=180, y=470, height=60, width=60)
    button_single_17.place(x=180, y=540, height=60, width=60)

    button_triple_16.place(x=240, y=400, height=60, width=60)
    button_double_16.place(x=240, y=470, height=60, width=60)
    button_single_16.place(x=240, y=540, height=60, width=60)
    button_triple_15.place(x=300, y=400, height=60, width=60)
    button_double_15.place(x=300, y=470, height=60, width=60)
    button_single_15.place(x=300, y=540, height=60, width=60)
    button_triple_14.place(x=360, y=400, height=60, width=60)
    button_double_14.place(x=360, y=470, height=60, width=60)
    button_single_14.place(x=360, y=540, height=60, width=60)
    button_triple_13.place(x=420, y=400, height=60, width=60)
    button_double_13.place(x=420, y=470, height=60, width=60)
    button_single_13.place(x=420, y=540, height=60, width=60)

    button_triple_12.place(x=480, y=400, height=60, width=60)
    button_double_12.place(x=480, y=470, height=60, width=60)
    button_single_12.place(x=480, y=540, height=60, width=60)
    button_triple_11.place(x=540, y=400, height=60, width=60)
    button_double_11.place(x=540, y=470, height=60, width=60)
    button_single_11.place(x=540, y=540, height=60, width=60)
    button_triple_10.place(x=600, y=400, height=60, width=60)
    button_double_10.place(x=600, y=470, height=60, width=60)
    button_single_10.place(x=600, y=540, height=60, width=60)
    button_triple_9.place(x=660, y=400, height=60, width=60)
    button_double_9.place(x=660, y=470, height=60, width=60)
    button_single_9.place(x=660, y=540, height=60, width=60)

    button_triple_8.place(x=720, y=400, height=60, width=60)
    button_double_8.place(x=720, y=470, height=60, width=60)
    button_single_8.place(x=720, y=540, height=60, width=60)
    button_triple_7.place(x=780, y=400, height=60, width=60)
    button_double_7.place(x=780, y=470, height=60, width=60)
    button_single_7.place(x=780, y=540, height=60, width=60)
    button_triple_6.place(x=840, y=400, height=60, width=60)
    button_double_6.place(x=840, y=470, height=60, width=60)
    button_single_6.place(x=840, y=540, height=60, width=60)
    button_triple_5.place(x=900, y=400, height=60, width=60)
    button_double_5.place(x=900, y=470, height=60, width=60)
    button_single_5.place(x=900, y=540, height=60, width=60)

    enable_throw_buttons2()


def enable_throw_buttons2():
    """
    This function enables button (0 to T20)
    :return:
    """

    button_triple_4.place(x=960, y=400, height=60, width=60)
    button_double_4.place(x=960, y=470, height=60, width=60)
    button_single_4.place(x=960, y=540, height=60, width=60)
    button_triple_3.place(x=1020, y=400, height=60, width=60)
    button_double_3.place(x=1020, y=470, height=60, width=60)
    button_single_3.place(x=1020, y=540, height=60, width=60)
    button_triple_2.place(x=1080, y=400, height=60, width=60)
    button_double_2.place(x=1080, y=470, height=60, width=60)
    button_single_2.place(x=1080, y=540, height=60, width=60)
    button_triple_1.place(x=1140, y=400, height=60, width=60)
    button_double_1.place(x=1140, y=470, height=60, width=60)
    button_single_1.place(x=1140, y=540, height=60, width=60)

    button_single_bull.place(x=1210, y=400, height=60, width=60)
    button_bull.place(x=1210, y=470, height=60, width=60)
    button_0.place(x=1210, y=540, height=60, width=60)


if __name__ == "__main__":
    # configure the window to generate
    gui = Tk()
    gui.geometry('1275x645')
    gui.resizable(width=False, height=False)
    gui.title("Darts counter - User Interface")
    gui.configure(background='grey')

    # define the exit - button
    exit_button = Button(gui, text="End game", command=button_exit, fg="black", bg="lightgreen",
                         font=('Arial', 10, 'bold'))
    exit_button.place(x=1175, y=0, height=80, width=100)

    # label for introducing and welcome
    label_welcome = Label(gui, text="Welcome to the darts - counter!\n"
                                    " Please select the number of players "
                                    "and the starting points.", bg="grey",
                          font=('Arial', 14))
    label_welcome.place(x=337.5, y=20, height=50, width=600)

    # label and buttons for selecting the number of players
    label_number_players = Label(gui, text="2", fg="black", font=('Arial', 13, 'bold'))

    button_plus_number_players = Button(gui, text="+", fg="black", bg="lightgreen",
                                        font=('Arial', 13, 'bold'), command=plus_player_numbers)
    button_minus_number_players = Button(gui, text="-", fg="black", bg="red",
                                         font=('Arial', 13, 'bold'), command=minus_player_numbers)

    label_number_players.place(x=587.5, y=200, height=30, width=100)
    button_minus_number_players.place(x=557.5, y=200, height=30, width=30)
    button_plus_number_players.place(x=687.5, y=200, height=30, width=30)

    # text input for 4 player names
    input_name1 = Entry(gui, bd=1, font=('Arial', 13))
    input_name2 = Entry(gui, bd=1, font=('Arial', 13))
    input_name3 = Entry(gui, bd=1, font=('Arial', 13))
    input_name4 = Entry(gui, bd=1, font=('Arial', 13))

    # default: 2 input fields are enabled
    input_name1.place(x=583.5, y=280, height=30, width=110)
    input_name2.place(x=583.5, y=330, height=30, width=110)

    # button for checking names and starting the game
    button_continue = Button(gui, text="Continue", bg="lightgreen", fg="black",
                             font=('Arial', 13, 'bold'), command=check_names)
    button_continue.place(x=750, y=120, height=30, width=100)

    # label and button for selecting the starting points (301 or 501)
    label_switch_starting_points = Label(gui, text="501", fg="black", font=('Arial', 13, 'bold'))
    button_inc_starting_points = Button(gui, text="+", fg="black",
                                        bg="lightgreen", font=('Arial', 10),
                                        command=increment_starting_points)
    button_dec_starting_points = Button(gui, text="-", fg="black", bg="red", font=('Arial', 10),
                                        command=decrement_starting_points)

    label_switch_starting_points.place(x=587.5, y=120, height=30, width=100)
    button_dec_starting_points.place(x=557.5, y=120, height=30, width=30)
    button_inc_starting_points.place(x=687.5, y=120, height=30, width=30)

    # labels for 4 players
    label_player_1_name = Label(gui, text="", fg="black", font=('Arial', 13, 'bold'))
    label_player_2_name = Label(gui, text="", fg="black", font=('Arial', 13, 'bold'))
    label_player_3_name = Label(gui, text="", fg="black", font=('Arial', 13, 'bold'))
    label_player_4_name = Label(gui, text="", fg="black", font=('Arial', 13, 'bold'))

    # labels for game score for 4 players
    label_1_score = Label(gui, text="501", fg="black", bg="white", font=('Arial', 13, 'bold'))
    label_2_score = Label(gui, text="501", fg="black", bg="white", font=('Arial', 13, 'bold'))
    label_3_score = Label(gui, text="501", fg="black", bg="white", font=('Arial', 13, 'bold'))
    label_4_score = Label(gui, text="501", fg="black", bg="white", font=('Arial', 13, 'bold'))

    # start - button und stop - button
    button_start_game = Button(gui, text="Start", bd=4, fg="black", bg="yellow", font=('Arial', 11),
                               command=start_game)
    # button_start_game.place(x=850, y=90, height=30, width=100)

    button_stop_game = Button(gui, text="Stop", bd=4, fg="black", bg="red", font=('Arial', 11),
                              command=start_game)

    # next - button
    next_button = Button(gui, text="Next", bd=4, fg="black", bg="yellow", font=('Arial', 11),
                         command=next_button)

    # textarea for scoring points which shall be decremented
    label_dart_score = Label(gui, text="", bd=4, font=('Arial', 13))
    button_dart_score = Button(gui, text="Add", bd=4, fg="black",
                               bg="lightgreen", font=('Arial', 10),
                               command=add)

    # label, which will be invisible, for calculating scores at the end
    label_invisible = Label(gui, text="", bd=4, font=('Arial', 13))

    # create Label for 1,2 and 3 Darts and count down button
    label_first_dart = Label(gui, text="1", bd=4, bg="yellow", font=('Arial', 13))
    label_second_dart = Label(gui, text="2", bd=4, bg="white", font=('Arial', 13))
    label_third_dart = Label(gui, text="3", bd=4, bg="white", font=('Arial', 13))

    zwischen_label = Label(gui, text="0", bd=4, bg="yellow", font=('Arial', 13))
    count_down_button = Button(gui, text="Count down", bd=4, fg="black",
                               bg="lightgreen", font=('Arial', 10),
                               command=count_down)

    # calculate kpis button
    button_create_excel = Button(gui, text="Calculate Score", bd=4, fg="black",
                                 bg="lightblue", font=('Arial', 11),
                                 command=create_excel)
    # button_create_excel.place(x=1125, y=250, height=80, width=150)

    # create all triple, double and single buttons
    button_triple_20 = Button(gui, text="T20", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=t20)
    button_single_20 = Button(gui, text="S20", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=s20)
    button_double_20 = Button(gui, text="D20", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=d20)

    button_triple_19 = Button(gui, text="T19", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=t19)
    button_single_19 = Button(gui, text="S19", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=s19)
    button_double_19 = Button(gui, text="D19", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=d19)

    button_triple_18 = Button(gui, text="T18", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=t18)
    button_single_18 = Button(gui, text="S18", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=s18)
    button_double_18 = Button(gui, text="D18", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=d18)

    button_triple_17 = Button(gui, text="T17", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=t17)
    button_single_17 = Button(gui, text="S17", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=s17)
    button_double_17 = Button(gui, text="D17", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=d17)

    button_triple_16 = Button(gui, text="T16", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=t16)
    button_single_16 = Button(gui, text="S16", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=s16)
    button_double_16 = Button(gui, text="D16", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=d16)
    button_triple_15 = Button(gui, text="T15", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=t15)
    button_single_15 = Button(gui, text="S15", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=s15)
    button_double_15 = Button(gui, text="D15", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=d15)
    button_triple_14 = Button(gui, text="T14", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=t14)
    button_single_14 = Button(gui, text="S14", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=s14)
    button_double_14 = Button(gui, text="D14", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=d14)

    button_triple_13 = Button(gui, text="T13", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=t13)
    button_single_13 = Button(gui, text="S13", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=s13)
    button_double_13 = Button(gui, text="D13", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=d13)
    button_triple_12 = Button(gui, text="T12", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=t12)
    button_single_12 = Button(gui, text="S12", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=s12)
    button_double_12 = Button(gui, text="D12", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=d12)
    button_triple_11 = Button(gui, text="T11", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=t11)
    button_single_11 = Button(gui, text="S11", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=s11)
    button_double_11 = Button(gui, text="D11", bd=4, fg="black", bg="green", font=('Arial', 14),
                              command=d11)

    button_triple_10 = Button(gui, text="T10", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=t10)
    button_single_10 = Button(gui, text="S10", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=s10)
    button_double_10 = Button(gui, text="D10", bd=4, fg="black", bg="red", font=('Arial', 14),
                              command=d10)
    button_triple_9 = Button(gui, text="T9", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=t_9)
    button_single_9 = Button(gui, text="S9", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=s_9)
    button_double_9 = Button(gui, text="D9", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=d_9)
    button_triple_8 = Button(gui, text="T8", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=t_8)
    button_single_8 = Button(gui, text="S8", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=s_8)
    button_double_8 = Button(gui, text="D8", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=d_8)

    button_triple_7 = Button(gui, text="T7", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=t_7)
    button_single_7 = Button(gui, text="S7", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=s_7)
    button_double_7 = Button(gui, text="D7", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=d_7)
    button_triple_6 = Button(gui, text="T6", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=t_6)
    button_single_6 = Button(gui, text="S6", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=s_6)
    button_double_6 = Button(gui, text="D6", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=d_6)
    button_triple_5 = Button(gui, text="T5", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=t_5)
    button_single_5 = Button(gui, text="S5", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=s_5)
    button_double_5 = Button(gui, text="D5", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=d_5)

    button_triple_4 = Button(gui, text="T4", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=t_4)
    button_single_4 = Button(gui, text="S4", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=s_4)
    button_double_4 = Button(gui, text="D4", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=d_4)
    button_triple_3 = Button(gui, text="T3", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=t_3)
    button_single_3 = Button(gui, text="S3", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=s_3)
    button_double_3 = Button(gui, text="D3", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=d_3)
    button_triple_2 = Button(gui, text="T2", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=t_2)
    button_single_2 = Button(gui, text="S2", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=s_2)
    button_double_2 = Button(gui, text="D2", bd=4, fg="black", bg="red", font=('Arial', 14),
                             command=d_2)

    button_triple_1 = Button(gui, text="T1", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=t_1)
    button_single_1 = Button(gui, text="S1", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=s_1)
    button_double_1 = Button(gui, text="D1", bd=4, fg="black", bg="green", font=('Arial', 14),
                             command=d_1)
    button_single_bull = Button(gui, text="25", bd=4, fg="black", bg="green", font=('Arial', 14),
                                command=single_bull)
    button_bull = Button(gui, text="BULL", bd=4, fg="black", bg="red", font=('Arial', 14),
                         command=bull)
    button_0 = Button(gui, text="0", bd=4, fg="black", bg="green", font=('Arial', 14),
                      command=null)

    # reset - button
    reset_button = Button(gui, text="Go back", bd=4, fg="black", bg="red", font=('Arial', 11),
                          command=reset)

    gui.mainloop()
